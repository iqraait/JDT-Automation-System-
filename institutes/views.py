import os
import zipfile
from io import BytesIO
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from academics.models import (
    Course, FormField, FormSection, CourseCategory, CourseSubCategory, ApplicationFeeType,
    ExamSubject, Class, Subject, NoticeBoard, Timetable, AcademicResult, StudentDocument,
    ClassYear, FeeCategoryMaster, FeeType, FeeStructure, FeeHead, StudentFeePayment
)
from applications.models import Application, ApplicationFieldValue, FeeCategory, Admission, TrashedStudent
from .models import Institute, AcademicYear
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import json
import datetime
from core.utils import generate_application_pdf


User = get_user_model()


def get_current_institute(request):
    """
    Helper function to retrieve the active institute for the logged-in user.
    Respects active_institute_id stored in session when switching institutes.
    Falls back to request.user.institute or Institute.objects.first().
    """
    if not hasattr(request, 'user') or not request.user.is_authenticated:
        return None
    if hasattr(request.user, 'get_active_institute'):
        inst = request.user.get_active_institute(request)
        if inst:
            return inst
    inst = getattr(request.user, 'institute', None)
    if not inst and (getattr(request.user, 'is_staff', False) or getattr(request.user, 'is_superuser', False)):
        from .models import Institute
        inst = Institute.objects.first()
    return inst




# =========================
# EMAIL CONFIGURATION (SMTP)
# =========================
# Please update your SMTP settings in settings.py or here
# You can use Gmail, Outlook, or any other SMTP provider.
# Example for Gmail:
# EMAIL_HOST = 'smtp.gmail.com'
# EMAIL_PORT = 587
# EMAIL_USE_TLS = True
# EMAIL_HOST_USER = 'your-email@gmail.com'
# EMAIL_HOST_PASSWORD = 'your-app-password'

from django.template.loader import render_to_string
from django.utils.html import strip_tags

def send_admission_email(admission):
    """Sends a selection/admission memo to the student email."""
    try:
        student = admission.application.student
        subject = f"Admission Selected - {admission.application.course.name}"
        
        # REQUIREMENT: Domain Integration & IP Fallback
        domain = settings.ALLOWED_HOSTS[0] if settings.ALLOWED_HOSTS and settings.ALLOWED_HOSTS[0] != '*' else '15.207.187.228'
        protocol = 'https' if 'jdtislam' in domain else 'http'
        # If accessing via IP or localhost, include the :8000 port
        port = ':8000' if not 'jdtislam' in domain or '127.0.0.1' in domain else ''
        login_url = f"{protocol}://{domain}{port}/accounts/login/"
        
        html_message = render_to_string('emails/status_update.html', {
            'student_name': student.first_name or student.username,
            'course_name': admission.application.course.name,
            'status_display': 'Selected for Admission',
            'status_slug': 'selected',
            'remarks': f"Registered Number: {admission.register_number}",
            'login_url': login_url
        })
        plain_message = strip_tags(html_message)
        
        send_mail(
            subject, 
            plain_message, 
            settings.EMAIL_HOST_USER, 
            [student.email], 
            html_message=html_message,
            fail_silently=True
        )
    except Exception as e:
        print(f"Email failed: {e}")

def send_status_email(application, new_status):
    """Sends a professional HTML status update email to the student."""
    try:
        student = application.student
        if not student or not student.email:
            print(f"Skipping email for Application #{application.id}: No student email found.")
            return
        
        status_map = {'pending': 'Under Audit', 'selected': 'Verified', 'rejected': 'Rejected', 'hold': 'On Hold'}
        status_display = status_map.get(new_status, str(new_status).title())
        
        subject = f"Update: Application for {application.course.name} is {status_display}"
        
        # Use a safe fallback for the login URL
        try:
            domain = settings.ALLOWED_HOSTS[0] if settings.ALLOWED_HOSTS else '127.0.0.1'
        except:
            domain = '127.0.0.1'
            
        # REQUIREMENT: Domain Integration & IP Fallback
        protocol = 'https' if 'jdtislam' in domain else 'http'
        port = ':8000' if not 'jdtislam' in domain or '127.0.0.1' in domain else ''
        login_url = f"{protocol}://{domain}{port}/accounts/login/"
        
        html_message = render_to_string('emails/status_update.html', {
            'student_name': student.first_name or student.username,
            'course_name': application.course.name if application.course else "Course",
            'status_display': status_display,
            'status_slug': new_status,
            'remarks': application.remarks or "",
            'login_url': login_url
        })
        plain_message = strip_tags(html_message)
        
        send_mail(
            subject, 
            plain_message, 
            settings.EMAIL_HOST_USER, 
            [student.email], 
            html_message=html_message,
            fail_silently=False  # Changed to False to see errors
        )
        print(f"Status email sent to {student.email} for Application #{application.id}")
    except Exception as e:
        print(f"CRITICAL: Status email failed for App #{application.id}: {e}")

def send_admission_status_email(admission, new_status):
    """Sends a professional HTML status update email for already admitted students."""
    try:
        student = admission.application.student
        if not student.email: return
        
        status_map = dict(Admission.ADMISSION_STATUS)
        status_display = status_map.get(new_status, new_status)
        
        subject = f"Student Record Alert: {status_display}"
        # REQUIREMENT: Domain Integration & IP Fallback
        domain = settings.ALLOWED_HOSTS[0] if settings.ALLOWED_HOSTS and settings.ALLOWED_HOSTS[0] != '*' else '15.207.187.228'
        protocol = 'https' if 'jdtislam' in domain else 'http'
        port = ':8000' if not 'jdtislam' in domain or '127.0.0.1' in domain else ''
        login_url = f"{protocol}://{domain}{port}/accounts/login/"
        
        html_message = render_to_string('emails/status_update.html', {
            'student_name': student.first_name or student.username,
            'course_name': admission.application.course.name,
            'status_display': status_display,
            'status_slug': new_status if new_status != 'active' else 'selected',
            'remarks': admission.status_reason,
            'login_url': login_url
        })
        plain_message = strip_tags(html_message)
        
        send_mail(
            subject, 
            plain_message, 
            settings.EMAIL_HOST_USER, 
            [student.email], 
            html_message=html_message,
            fail_silently=True
        )
    except Exception as e:
        print(f"Admission status email failed: {e}")


# =========================
# HELPER FOR ALLOTMENT MEMO
# =========================
def get_allotment_memo_defaults(app):
    dob = None
    place = None
    place_fields = ['district', 'city', 'place', 'address', 'location']
    
    for fv in app.field_values.all():
        label = (fv.field.label if fv.field else fv.field_label or "").lower()
        if ('dob' in label or 'date of birth' in label) and not dob:
            dob = fv.value
        for pf in place_fields:
            if pf in label and not place:
                place = fv.value
                break
                
    if getattr(app, 'application_no', None):
        app_no = app.application_no
    elif getattr(app, 'form_no', None):
        app_no = app.form_no
    else:
        app_no_val = None
        for fv in app.field_values.all():
            lbl = (fv.field.label if fv.field else fv.field_label or "").lower()
            if 'application' in lbl or 'form no' in lbl or 'app no' in lbl:
                app_no_val = fv.value
                break
        app_no = app_no_val or str(app.id)

    admission = Admission.objects.filter(application=app).first()
    quota_val = None
    if admission and admission.admission_quota:
        quota_val = admission.admission_quota
    else:
        for fv in app.field_values.all():
            lbl = (fv.field.label if fv.field else fv.field_label or "").lower()
            if 'quota' in lbl or 'category' in lbl:
                quota_val = fv.value
                break
    if not quota_val:
        if app.course and app.course.category:
            quota_val = app.course.category.name
        else:
            quota_val = 'Management'

    return {
        'app_no': app_no,
        'student_name': get_student_name(app),
        'dob': dob or 'N/A',
        'place': place or 'N/A',
        'quota': quota_val,
    }


# =========================
# ADMISSION LIST
# =========================
@login_required
def admission_list(request):
    institute = request.user.institute
    
    # Only "Selected" applications
    applications = Application.objects.filter(institute=institute, status='selected').select_related(
        'student', 'course', 'academic_year', 'course__category'
    ).order_by('-id')
    
    # Filters
    form_id = request.GET.get('form_id')
    name = request.GET.get('name')
    admission_year = request.GET.get('year')
    course_id = request.GET.get('course_id')
    status_filter = request.GET.get('status')
    
    if form_id:
        applications = applications.filter(id=form_id)
        
    if admission_year:
        applications = applications.filter(academic_year_id=admission_year)

    if course_id:
        applications = applications.filter(course_id=course_id)

    if status_filter == 'enrolled':
        applications = applications.filter(admission__isnull=False)
    elif status_filter == 'pending':
        applications = applications.filter(admission__isnull=True)

    from django.db.models import Q
    if name:
        name_clean = name.strip()
        applications = applications.filter(
            Q(student__first_name__icontains=name_clean) |
            Q(student__last_name__icontains=name_clean) |
            Q(student__username__icontains=name_clean)
        )

    # Get active academic years related to selected institute
    years = AcademicYear.objects.filter(institute=institute, is_active=True)
    courses = Course.objects.filter(institute=institute)

    # Paginate Applications QuerySet FIRST
    from django.core.paginator import Paginator
    paginator = Paginator(applications, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Fetch page objects and prefetch field_values ONLY for these page items (max 20)
    page_app_ids = [app.id for app in page_obj.object_list]
    page_apps_qs = Application.objects.filter(id__in=page_app_ids).select_related(
        'student', 'course', 'academic_year', 'course__category'
    ).prefetch_related(
        'field_values', 'field_values__field', 'field_values__field__section', 'admission'
    )
    page_apps_dict = {app.id: app for app in page_apps_qs}

    # Calculate ranks for displayed applications without hitting SQLite limits (chunk_size=400)
    course_year_groups = set()
    for app in page_obj.object_list:
        if app.course_id and app.academic_year_id:
            course_year_groups.add((app.course_id, app.academic_year_id))
            
    rank_map = {}
    for cid, ayid in course_year_groups:
        group_app_ids = list(Application.objects.filter(
            institute=institute,
            status='selected',
            payment__status='success',
            course_id=cid,
            academic_year_id=ayid
        ).values_list('id', flat=True))

        ranked_list = []
        chunk_size = 400
        for i in range(0, len(group_app_ids), chunk_size):
            chunk_ids = group_app_ids[i:i + chunk_size]
            group_apps = Application.objects.filter(id__in=chunk_ids).select_related(
                'student', 'course', 'academic_year'
            ).prefetch_related(
                'field_values', 'field_values__field', 'field_values__field__section'
            )
            for g_app in group_apps:
                total, percentage, main_mark, sub_mark, max_total, qualified_total = calculate_total_and_percentage(g_app)
                ranked_list.append({
                    "app_id": g_app.id,
                    "percentage": percentage,
                    "main_mark": main_mark,
                    "sub_mark": sub_mark
                })

        ranked_list.sort(key=lambda x: (x['percentage'], x['main_mark'], x['sub_mark']), reverse=True)
        for idx, item in enumerate(ranked_list, start=1):
            rank_map[item['app_id']] = idx

    processed_apps = []
    for app_item in page_obj.object_list:
        app = page_apps_dict.get(app_item.id, app_item)
        std_name = get_student_name(app)
            
        memo_def = get_allotment_memo_defaults(app)

        processed_apps.append({
            "form_no": app.id,
            "name": std_name,
            "student_first_name": app.student.first_name if app.student.first_name else std_name,
            "course": app.course,
            "academic_year": app.academic_year,
            "status": app.status,
            "is_registered": hasattr(app, 'admission'),
            "category": app.course.category.name if (app.course and app.course.category) else "Uncategorized",
            "rank": rank_map.get(app.id, 'N/A'),
            "app_no": memo_def['app_no'],
            "dob": memo_def['dob'],
            "place": memo_def['place'],
            "quota": memo_def['quota'],
        })

    # Wrap processed apps into a paginator page object structure for the template
    page_obj.object_list = processed_apps

    return render(request, 'institute/admission_list.html', {
        'applications': page_obj,
        'page_obj': page_obj,
        'years': years,
        'courses': courses,
        'selected_year': admission_year,
        'selected_course_id': course_id,
        'selected_status': status_filter,
        'selected_name': name,
        'selected_form_id': form_id
    })


# =========================
# ALLOTMENT MEMO
# =========================
@login_required
def generate_allotment_memo(request, app_id):
    app = get_object_or_404(Application, id=app_id)
    institute = request.user.institute
    
    include_bank = request.GET.get('include_bank_details') in ['on', 'true', '1']
    memo_def = get_allotment_memo_defaults(app)
    
    context = {
        'phase_no': request.GET.get('phase_no', '1st Allotment'),
        'allotment_date': request.GET.get('allotment_date'),
        'app_no': request.GET.get('app_no') or memo_def['app_no'],
        'student_name': request.GET.get('student_name') or memo_def['student_name'],
        'dob': request.GET.get('dob') or memo_def['dob'],
        'quota': request.GET.get('quota') or memo_def['quota'],
        'place': request.GET.get('place') or memo_def['place'],
        'rank': request.GET.get('rank', ''),
        'reporting_time': request.GET.get('reporting_time', '10:00 AM'),
        'report_from': request.GET.get('report_from'),
        'report_to': request.GET.get('report_to'),
        'fee_details': request.GET.get('fee_details', ''),
        'include_bank_details': include_bank,
        'bank_name': request.GET.get('bank_name', 'Punjab National Bank'),
        'account_no': request.GET.get('account_no', '4909001300014441'),
        'ifsc_code': request.GET.get('ifsc_code', 'PUNB0788500'),
        'branch_name': request.GET.get('branch_name', 'Vellimadukunnu'),
        'app': app,
        'institute': institute,
    }
    
    return render(request, 'institute/allotment_memo.html', context)


@login_required
def send_allotment_memo_email(request, app_id):
    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string

    app = get_object_or_404(Application, id=app_id, institute=request.user.institute)
    student_email = app.student.email if app.student else None
    
    if not student_email:
        messages.error(request, "Student does not have a registered email address.")
        return redirect(request.META.get('HTTP_REFERER', '/institute/admission/'))

    include_bank = request.GET.get('include_bank_details') in ['on', 'true', '1']
    memo_def = get_allotment_memo_defaults(app)

    context = {
        'phase_no': request.GET.get('phase_no', '1st Allotment'),
        'allotment_date': request.GET.get('allotment_date'),
        'app_no': request.GET.get('app_no') or memo_def['app_no'],
        'student_name': request.GET.get('student_name') or memo_def['student_name'],
        'dob': request.GET.get('dob') or memo_def['dob'],
        'quota': request.GET.get('quota') or memo_def['quota'],
        'place': request.GET.get('place') or memo_def['place'],
        'rank': request.GET.get('rank', ''),
        'reporting_time': request.GET.get('reporting_time', '10:00 AM'),
        'report_from': request.GET.get('report_from'),
        'report_to': request.GET.get('report_to'),
        'fee_details': request.GET.get('fee_details', ''),
        'include_bank_details': include_bank,
        'bank_name': request.GET.get('bank_name', 'Punjab National Bank'),
        'account_no': request.GET.get('account_no', '4909001300014441'),
        'ifsc_code': request.GET.get('ifsc_code', 'PUNB0788500'),
        'branch_name': request.GET.get('branch_name', 'Vellimadukunnu'),
        'app': app,
        'institute': request.user.institute,
        'is_email': True,
    }

    html_content = render_to_string('institute/allotment_memo.html', context)
    student_display_name = context['student_name'] or app.display_name
    subject = f"Allotment Memo - {student_display_name} ({app.course.name if app.course else ''})"
    
    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=f"Dear {student_display_name},\n\nPlease find your Allotment Memo details for {app.course.name if app.course else ''}.\n\nRegards,\n{request.user.institute.name}",
            from_email=None,
            to=[student_email]
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send()
        messages.success(request, f"Allotment Memo successfully emailed to {student_email}")
    except Exception as e:
        messages.error(request, f"Failed to send email: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', '/institute/admission/'))


# =========================
# REGISTER STUDENT
# =========================
@login_required
def register_student(request, app_id):
    app = get_object_or_404(Application, id=app_id)
    course = app.course
    
    if request.method == 'POST':
        # 1. Save Admission Record
        registration_id = request.POST.get('registration_id')
        student_email = request.POST.get('student_email')
        date_of_join = request.POST.get('date_of_join')
        admission_quota = request.POST.get('admission_quota', 'Merit')
        fee_cat_id = request.POST.get('fee_category_id')
        joining_period_id = request.POST.get('joining_period_id')
        calculated_fee = request.POST.get('calculated_fee')
        
        # Update student email if changed
        if student_email and app.student.email != student_email:
            app.student.email = student_email
            app.student.save()
        discount_amount = request.POST.get('discount_amount', 0) or 0
        discount_reason = request.POST.get('discount_reason')
        final_fee = request.POST.get('final_fee')
        selected_course_id = request.POST.get('course_id')
        
        if selected_course_id:
            course = get_object_or_404(Course, id=selected_course_id)
            if app.course != course:
                app.course = course
                app.save()
        
        # Guardian
        care_of = request.POST.get('care_of', '') or ''
        guardian_name = request.POST.get('guardian_name', '') or ''
        guardian_mobile = request.POST.get('guardian_mobile', '') or ''
        relationship = request.POST.get('relationship', '') or ''
        guardian_address = request.POST.get('guardian_address', '') or ''
        
        if Admission.objects.filter(registration_id=registration_id).exclude(application=app).exists():
            messages.error(request, "Registration ID already exists.")
            return redirect(request.path)

        # Fix: Convert date string to date object
        doj_obj = datetime.datetime.strptime(date_of_join, '%Y-%m-%d').date() if date_of_join else datetime.date.today()

        fee_cat_master = None
        app_fee_cat = None
        if fee_cat_id:
            fee_cat_master = FeeCategoryMaster.objects.filter(id=fee_cat_id).first()
            if fee_cat_master:
                app_fee_cat = FeeCategory.objects.filter(course=course, name__iexact=fee_cat_master.name).first()

        # Check if Admission already exists
        adm = Admission.objects.filter(application=app).first()
        if adm:
            # Update existing Admission record
            adm.registration_id = registration_id
            adm.admission_quota = admission_quota
            adm.date_of_join = doj_obj
            adm.selected_course = course
            adm.fee_category = app_fee_cat
            adm.assigned_fee_category = fee_cat_master
            adm.joining_period_id = joining_period_id if joining_period_id else None
            adm.calculated_fee = calculated_fee if calculated_fee else 0.00
            adm.discount_amount = discount_amount if discount_amount else 0.00
            adm.discount_reason = discount_reason
            adm.final_fee = final_fee if final_fee else 0.00
            adm.care_of = care_of
            adm.guardian_name = guardian_name
            adm.guardian_mobile = guardian_mobile
            adm.relationship = relationship
            adm.guardian_address = guardian_address
            adm.assigned_class_id = request.POST.get('assigned_class_id') if request.POST.get('assigned_class_id') else None
            adm.assigned_class_year_id = request.POST.get('assigned_class_year_id') if request.POST.get('assigned_class_year_id') else None
            adm.save()
        else:
            # Create new Admission record
            adm = Admission.objects.create(
                application=app,
                registration_id=registration_id,
                admission_quota=admission_quota,
                date_of_join=doj_obj,
                selected_course=course,
                fee_category=app_fee_cat,
                assigned_fee_category=fee_cat_master,
                joining_period_id=joining_period_id if joining_period_id else None,
                calculated_fee=calculated_fee if calculated_fee else 0.00,
                discount_amount=discount_amount if discount_amount else 0.00,
                discount_reason=discount_reason,
                final_fee=final_fee if final_fee else 0.00,
                care_of=care_of,
                guardian_name=guardian_name,
                guardian_mobile=guardian_mobile,
                relationship=relationship,
                guardian_address=guardian_address,
                assigned_class_id=request.POST.get('assigned_class_id') if request.POST.get('assigned_class_id') else None,
                assigned_class_year_id=request.POST.get('assigned_class_year_id') if request.POST.get('assigned_class_year_id') else None
            )

        # 2. Save Dynamic Form Fields
        fields = FormField.objects.filter(form=course.form)
        for field in fields:
            key = f"field_{field.id}"
            if field.field_type == 'file':
                file_obj = request.FILES.get(key)
                if file_obj:
                    # FIX: Use FileSystemStorage to save file on disk
                    from django.core.files.storage import FileSystemStorage
                    fs = FileSystemStorage()
                    filename = fs.save(file_obj.name, file_obj)

                    # Handle potential duplicate values safely
                    val_obj = ApplicationFieldValue.objects.filter(application=app, field=field).first()
                    if not val_obj:
                        val_obj = ApplicationFieldValue.objects.create(
                            application=app, 
                            field=field,
                            field_label=field.label,
                            field_type=field.field_type
                        )
                    
                    val_obj.value = filename
                    val_obj.save()
            else:
                val = request.POST.get(key)
                if val is not None:
                    # FIX: Handle potential duplicate values safely
                    val_qs = ApplicationFieldValue.objects.filter(application=app, field=field)
                    if val_qs.exists():
                        val_qs.update(value=val, field_label=field.label, field_type=field.field_type)
                    else:
                        ApplicationFieldValue.objects.create(
                            application=app, 
                            field=field, 
                            field_label=field.label,
                            field_type=field.field_type,
                            value=val
                        )
        
        # 3. Save Qualifying Exam Marks
        # Only target fields in the 'Qualifying Examination' section
        qe_field = FormField.objects.filter(
            form=course.form, 
            section__name__icontains="Qualifying Examination"
        ).first()
        if qe_field:
            # Delete old marks for this field before re-saving
            ApplicationFieldValue.objects.filter(application=app, field=qe_field, value__contains=":").delete()
            for key in request.POST:
                if key.startswith("subject_"):
                    subject_name = key.replace("subject_", "").strip()
                    marks = request.POST.get(key)
                    max_marks = request.POST.get(f"max_{subject_name}", "100")
                    if marks:
                        # Clean up marks to avoid ".0" if they are integers, otherwise keep decimals
                        try:
                            m_val = float(marks)
                            marks_str = f"{int(m_val)}" if m_val == int(m_val) else f"{m_val}"
                        except:
                            marks_str = marks
                            
                        ApplicationFieldValue.objects.create(
                            application=app,
                            field=qe_field,
                            value=f"{subject_name}:{marks_str}:{max_marks}"
                        )
        
        
        messages.success(request, "this student is registered successfully")
        
        # SEND NOTIFICATION EMAIL
        send_admission_email(adm)
        
        return redirect('student_list')

    # GET Request
    student_name = get_student_name(app)
    
    # Auto-fill Mobile logic
    student_mobile = app.student.mobile_number
    if not student_mobile:
        for v in app.field_values.all():
            lbl = (v.field.label if v.field else v.field_label or "").lower()
            if "phone" in lbl or "mobile" in lbl or "contact" in lbl:
                student_mobile = v.value
                break
    
    # Check if Admission already exists
    adm = Admission.objects.filter(application=app).first()
    
    # Registration ID should be blank so they enter it manually, unless already registered
    registration_id = adm.registration_id if adm else ""
        
    fee_categories = FeeCategoryMaster.objects.filter(is_active=True)
    fee_cats_json = []
    for cat in fee_categories:
        app_fee_cat = FeeCategory.objects.filter(course=course, name__iexact=cat.name).first()
        fee_cats_json.append({
            'id': cat.id, 'name': cat.name, 'total': float(app_fee_cat.total_fee) if app_fee_cat else 0.0, 
            'breakdown': app_fee_cat.breakdown if app_fee_cat else [], 'type': 'semester'
        })
    
    # Fetch Dynamic Fields for the specific course form
    form_fields = FormField.objects.filter(form=course.form).order_by('section__order', 'order')
    sections = {}
    for f in form_fields:
        if f.section not in sections:
            sections[f.section] = []
        sections[f.section].append(f)
    
    # Fetch existing values for these fields
    field_values_by_id = {}
    field_values_by_label = {}
    for v in app.field_values.all().order_by('id'):
        val_str = str(v.value or "")
        if ":" not in val_str: 
            if v.field_id:
                field_values_by_id[v.field_id] = v.value
            lbl = (v.field.label if v.field else v.field_label or "").lower()
            if lbl:
                field_values_by_label[lbl] = v.value
    
    for f in form_fields:
        f.current_value = field_values_by_id.get(f.id, "")
        if not f.current_value:
            f.current_value = field_values_by_label.get(f.label.lower(), "")
            
        # FIX: Ensure Full Name shows student name, not corrupted subject marks
        if f.label == "Full Name" and (not f.current_value or ":" in str(f.current_value)):
            f.current_value = app.student.first_name
        f.value = f.current_value  # Support templates using .value or .current_value

    # Fetch marks for display
    subjects = []
    latest_subjects = {}
    for v in app.field_values.all().order_by('-id'):
        label_lower = (v.field.label if v.field else v.field_label or "").lower()
        section_lower = (v.field.section.name if v.field and v.field.section else v.field_label or "").lower()
        is_qual_field = any(x in label_lower or x in section_lower for x in ["mark", "subject", "qualify", "exam"])
        if v.value and ":" in str(v.value) and is_qual_field and not getattr(v.field, 'is_photo', False) and not getattr(v.field, 'is_signature', False):
            parts = str(v.value).split(":")
            if len(parts) >= 2:
                name = parts[0].strip()
                marks = parts[1].strip()
                max_marks = parts[2].strip() if len(parts) >= 3 else "100"
                if name not in latest_subjects:
                    latest_subjects[name] = {'marks': marks, 'max': max_marks}
    for name, data in latest_subjects.items():
        subjects.append({"name": name, "marks": data['marks'], "max": data['max']})
        
    subjects.reverse()

    # Get subcategories (labels) for the course category
    subcategories = course.category.subcategories.all() if course.category else []

    return render(request, 'institute/register_student.html', {
        'app': app,
        'adm': adm,
        'student_name': student_name,
        'registration_id': registration_id,
        'fee_categories': fee_categories,
        'fee_cats_json': json.dumps(fee_cats_json),
        'sections': sections,
        'subjects': subjects,
        'subcategories': subcategories,
        'student_mobile': student_mobile,
        'academic_years': AcademicYear.objects.filter(institute=app.institute),
        'courses': Course.objects.filter(institute=app.institute)
    })


from django.http import JsonResponse

# ✅ AJAX: LOAD SUBCATEGORIES
def load_subcategories(request):
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse([], safe=False)
    subcategories = CourseSubCategory.objects.filter(category__courses__id=course_id)
    data = [{'id': s.id, 'name': s.name} for s in subcategories]
    return JsonResponse(data, safe=False)


# ✅ AJAX: LOAD CLASSES
def load_classes(request):
    course_id = request.GET.get('course_id')
    academic_year_id = request.GET.get('academic_year_id')
    period_id = request.GET.get('period_id')
    category_id = request.GET.get('category_id') # New filter
    
    classes = Class.objects.filter(institute=request.user.institute)
    
    if course_id:
        classes = classes.filter(course_id=course_id)
    
    # Optional filters: If these exist on the class, they should match. 
    # But if the class has NO year/period assigned, it should still show up.
    # OR, if we want to show all classes for the course regardless of year/period for flexibility:
    
    # For now, let's keep it course-focused as requested by the user to "show my classes"
    # if academic_year_id:
    #     classes = classes.filter(academic_year_id=academic_year_id)
    # if period_id:
    #     classes = classes.filter(period_id=period_id)
        
    data = [{'id': c.id, 'name': c.name} for c in classes]
    return JsonResponse(data, safe=False)


# ✅ AJAX: LOAD CLASS YEARS
def load_class_years(request):
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse([], safe=False)
    class_years = ClassYear.objects.filter(class_obj_id=class_id, is_active=True)
    data = [{'id': cy.id, 'name': cy.name} for cy in class_years]
    return JsonResponse(data, safe=False)


# ✅ AJAX: LOAD FEE CATEGORIES
def load_fee_categories(request):
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse([], safe=False)
    course = get_object_or_404(Course, id=course_id)
    masters = FeeCategoryMaster.objects.filter(is_active=True)
    data = []
    for m in masters:
        app_cat = FeeCategory.objects.filter(course=course, name__iexact=m.name).first()
        data.append({
            'id': m.id,
            'name': m.name,
            'total_fee': float(app_cat.total_fee) if app_cat else 0.0,
            'breakdown': app_cat.breakdown if app_cat else []
        })
    return JsonResponse(data, safe=False)


# ✅ AJAX: LOAD FORM FIELDS
def load_form_fields(request):
    course_id = request.GET.get('course_id')
    if not course_id:
        return JsonResponse([], safe=False)
    course = get_object_or_404(Course, id=course_id)
    fields = FormField.objects.filter(form=course.form).order_by('section__order', 'order')
    
    data = []
    for f in fields:
        data.append({
            'id': f.id,
            'label': f.label,
            'type': f.field_type,
            'required': f.required,
            'section': f.section.name if f.section else 'General Info',
            'options': [{'value': o.value, 'text': o.display_text} for o in f.options.all()] if f.field_type == 'select' else []
        })
    return JsonResponse(data, safe=False)


def load_exam_subjects(request):
    exam_id = request.GET.get('exam_id')
    if not exam_id:
        return JsonResponse([], safe=False)
        
    # Find the QualifyingExam object
    exam_obj = None
    if str(exam_id).isdigit():
        exam_obj = QualifyingExam.objects.filter(id=exam_id).first()
    
    if not exam_obj:
        # Try finding by name if it's not a numeric ID
        exam_obj = QualifyingExam.objects.filter(name__iexact=exam_id).first()

    if not exam_obj:
        return JsonResponse([], safe=False)
        
    subjects_data = []
    for sub in ExamSubject.objects.filter(exam=exam_obj):
        subjects_data.append({
            'exam_id': exam_obj.id,
            'exam_name': exam_obj.name,
            'subject_id': sub.id,
            'subject_name': sub.name,
            'max_marks': sub.max_marks,
            'pass_mark': sub.pass_mark
        })
            
    return JsonResponse(subjects_data, safe=False)


# =========================
# MANUAL REGISTRATION
# =========================
@login_required
def register_manual(request):
    if request.method == 'POST':
        # 1. Create a dummy / actual Student User
        student_name = request.POST.get('student_name')
        mobile = request.POST.get('mobile')
        email = request.POST.get('email')
        
        # Check if user exists
        user, created = User.objects.get_or_create(
            username=mobile,
            defaults={
                'email': email,
                'role': 'student',
                'first_name': student_name
            }
        )
        if created:
            user.set_password('jdt123') # Default password
            user.save()

        registration_id = request.POST.get('registration_id')
        if Admission.objects.filter(registration_id=registration_id).exists():
            messages.error(request, f"Registration ID {registration_id} is already in use. Please use a unique ID.")
            return redirect('register_manual')

        # 2. Create Application
        institute = request.user.institute
        course_id = request.POST.get('course_id')
        academic_year_id = request.POST.get('academic_year_id')
        
        application = Application.objects.create(
            student=user,
            institute=institute,
            academic_year_id=academic_year_id,
            course_id=course_id,
            status='selected' # Manual registration usually means they are already selected
        )

        # 3. Save Form Fields (Similar to apply_course)
        fields = FormField.objects.filter(form__course_id=course_id)
        for field in fields:
            key = f'field_{field.id}'
            if field.field_type == 'file':
                file_obj = request.FILES.get(key)
                if file_obj:
                    fs = FileSystemStorage()
                    filename = fs.save(file_obj.name, file_obj)
                    ApplicationFieldValue.objects.create(
                        application=application, 
                        field=field,
                        field_label=field.label,
                        field_type=field.field_type,
                        value=filename
                    )
            else:
                value = request.POST.get(key)
                if value:
                    ApplicationFieldValue.objects.create(
                        application=application, 
                        field=field,
                        field_label=field.label,
                        field_type=field.field_type,
                        value=value
                    )

        # 4. Save Subjects for Meriting/Ranking
        # Link subjects to the 'Qualifying Examination' field if it exists
        qe_field = FormField.objects.filter(form__course_id=course_id, section__name__icontains="Qualifying Examination").first()
        if not qe_field:
            qe_field = fields.first()
            
        if qe_field:
            for key in request.POST:
                if key.startswith("subject_"):
                    subject_name = key.replace("subject_", "").strip()
                    marks = request.POST.get(key)
                    if marks:
                        ApplicationFieldValue.objects.create(
                            application=application,
                            field=qe_field, 
                            value=f"{subject_name}:{marks}:100" # Default max 100 for manual entries
                        )

        # 5. Create Admission
        registration_id = request.POST.get('registration_id')
        date_of_join = request.POST.get('date_of_join')
        fee_cat_id = request.POST.get('fee_category_id')
        joining_period_id = request.POST.get('joining_period_id')
        calculated_fee = request.POST.get('calculated_fee')
        discount_amount = request.POST.get('discount_amount', 0) or 0
        discount_reason = request.POST.get('discount_reason')
        final_fee = request.POST.get('final_fee')
        
        admission_quota = request.POST.get('admission_quota', 'Merit')
        care_of = request.POST.get('care_of', '') or ''
        guardian_name = request.POST.get('guardian_name', '') or ''
        guardian_mobile = request.POST.get('guardian_mobile', '') or ''
        relationship = request.POST.get('relationship', '') or ''
        guardian_address = request.POST.get('guardian_address', '') or ''

        # Fix: Convert date string to date object
        doj_obj = datetime.datetime.strptime(date_of_join, '%Y-%m-%d').date() if date_of_join else datetime.date.today()

        fee_cat_master = None
        app_fee_cat = None
        if fee_cat_id:
            fee_cat_master = FeeCategoryMaster.objects.filter(id=fee_cat_id).first()
            if fee_cat_master:
                app_fee_cat = FeeCategory.objects.filter(course_id=course_id, name__iexact=fee_cat_master.name).first()

        adm = Admission.objects.create(
            application=application,
            registration_id=registration_id,
            admission_quota=admission_quota,
            date_of_join=doj_obj,
            selected_course_id=course_id,
            fee_category=app_fee_cat,
            assigned_fee_category=fee_cat_master,
            joining_period_id=joining_period_id if joining_period_id else None,
            assigned_class_id=request.POST.get('assigned_class_id') if request.POST.get('assigned_class_id') else None,
            assigned_class_year_id=request.POST.get('assigned_class_year_id') if request.POST.get('assigned_class_year_id') else None,
            calculated_fee=calculated_fee if calculated_fee else 0.00,
            discount_amount=discount_amount if discount_amount else 0.00,
            discount_reason=discount_reason,
            final_fee=final_fee if final_fee else 0.00,
            care_of=care_of,
            guardian_name=guardian_name,
            guardian_mobile=guardian_mobile,
            relationship=relationship,
            guardian_address=guardian_address
        )

        # Create Payment record with status 'success' for student register & dashboard integration
        Payment.objects.get_or_create(
            application=application,
            defaults={
                'amount': final_fee if final_fee else (app_fee_cat.total_fee if app_fee_cat else 0.00),
                'status': 'success',
                'payment_mode': 'CASH',
                'payment_date': doj_obj
            }
        )
        
        
        messages.success(request, "this student is registered successfully")
        
        # SEND NOTIFICATION EMAIL
        send_admission_email(adm)
        
        return redirect('student_list')
    institutes = [request.user.institute]
    academic_years = AcademicYear.objects.filter(institute=request.user.institute, is_active=True)
    courses = Course.objects.filter(institute=request.user.institute)
    
    fee_cats = []
    for cat in FeeCategoryMaster.objects.filter(is_active=True):
        fee_cats.append({
            'id': cat.id,
            'name': cat.name,
            'total_fee': 0.0,
            'breakdown': '[]'
        })

    return render(request, 'institute/register_manual.html', {
        'institutes': institutes,
        'academic_years': academic_years,
        'courses': courses,
        'fee_categories': fee_cats,
        # For manual registration, we might need to load form fields via AJAX (already handled in template)
    })


# =========================
# LOGIN & ROOT REDIRECT
# =========================
def root_institute_view(request):
    if request.user.is_authenticated and (getattr(request.user, 'role', '') == 'institute' or hasattr(request.user, 'institute')):
        return redirect('/institute/dashboard/')
    return redirect('/institute/login/')


def institute_login(request):
    if request.user.is_authenticated and getattr(request.user, 'role', '') == 'institute':
        return redirect('/institute/dashboard/')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:

            if user.role != 'institute':
                messages.error(request, "Not an institute account")
                return redirect('/institute/login/')

            if not hasattr(user, 'institute'):
                messages.error(request, "Institute not assigned")
                return redirect('/institute/login/')

            login(request, user)
            return redirect('/institute/dashboard/')

        else:
            messages.error(request, "Invalid Credentials")

    return render(request, 'institute/login.html')


# =========================
# GENERATE PDF
# =========================




# =========================
# DOWNLOAD PDF
# =========================
def download_application_zip(request, app_id):

    application = Application.objects.get(id=app_id)

    # 🔥 Create ZIP in memory
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:

        # ===== PDF =====
        pdf_buffer = BytesIO()
        generate_application_pdf(application, pdf_buffer)

        pdf_filename = f"application_{application.id}.pdf"
        zip_file.writestr(pdf_filename, pdf_buffer.getvalue())

        # ===== FILES =====
        for v in application.field_values.all():

            if v.field.field_type == "file" and v.value:

                file_path = os.path.join(settings.MEDIA_ROOT, str(v.value))
                if os.path.exists(file_path):
                     with open(file_path, 'rb') as f:
                         filename = os.path.basename(file_path)
                         # Add prefix to distinguish files
                         field_label = v.field.label if v.field else v.field_label
                         label_slug = field_label.replace(" ", "_").lower() if field_label else f"field_{v.id}"
                         zip_file.writestr(f"documents/{label_slug}_{filename}", f.read())


    zip_buffer.seek(0)

    student_name = get_student_name(application).replace(" ", "_")

    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={student_name}_Form_{application.id}.zip'

    return response


@login_required
def view_application(request, app_id):
    """
    Read-only view for the admission form.
    Accessible by the student who owns it or the institute staff.
    """
    if request.user.role == 'institute':
        # Ensure institute user is only viewing applications for their institute
        application = get_object_or_404(Application, id=app_id, institute=request.user.institute)
    else:
        # Students can only view their own applications
        application = get_object_or_404(Application, id=app_id, student=request.user)

    # 1. Fetch ALL field values first to avoid queryset issues
    field_values = list(application.field_values.select_related('field', 'field__section').all())
    
    # Identify photo and signature
    student_photo = None
    student_signature = None
    subject_marks = []
    total_obtained = 0
    total_max = 0

    # 1. Identify Qualifying Exam and Subjects Configuration
    from academics.models import QualifyingExam, ExamSubject
    exam_obj = None
    
    # Step A: Find the examination name from the form values
    exam_name_from_form = None
    for fv in field_values:
        label = (fv.field.label if fv.field else fv.field_label or "").lower()
        if ("exam" in label or "qualifying" in label) and "marks" not in label:
            val = str(fv.value).strip()
            # If it's a choice field, get the display text first
            if fv.field and fv.field.field_type in ['select', 'radio']:
                from academics.models import FieldOption
                opt = FieldOption.objects.filter(field=fv.field, value=val).first()
                if opt:
                    exam_name_from_form = opt.display_text
            
            # Fallback to the raw value if no option found
            if not exam_name_from_form:
                exam_name_from_form = val
            break

    # Step B: Resolve the QualifyingExam object based on the name we found
    subjects_config = {}
    if exam_name_from_form:
        # Try finding by name (this is more reliable than ID which might mismatch with FieldOptions)
        exam_obj = QualifyingExam.objects.filter(name__iexact=exam_name_from_form).first()
        if exam_obj:
            for s in ExamSubject.objects.filter(exam=exam_obj):
                subjects_config[s.name.lower().strip()] = s.max_marks
        
        # Fallback to ID if name lookup fails and it's numeric
        if not exam_obj and exam_name_from_form.isdigit():
            exam_obj = QualifyingExam.objects.filter(id=exam_name_from_form).first()
            if exam_obj:
                for s in ExamSubject.objects.filter(exam=exam_obj):
                    subjects_config[s.name.lower().strip()] = s.max_marks

    # 2. Process ALL field values to extract marks and media (including orphans/mislinked)
    processed_fv_ids = set()
    for fv in field_values:
        label_lower = (fv.field.label if fv.field else fv.field_label or "").lower()
        val_str = str(fv.value or "").strip()
        section_lower = (fv.field.section.name if fv.field and fv.field.section else fv.field_label or "").lower()
        is_qual_field = any(x in label_lower or x in section_lower for x in ["mark", "subject", "qualify", "exam"])

        # Identify subject marks (value containing ':')
        # Broad check: if it has a colon and the second part is numeric, it's likely a mark
        is_mark_format = ":" in val_str and len(val_str.split(":")) >= 2 and is_qual_field
        is_media_field = any(x in label_lower for x in ["photo", "signature", "sign"])
        
        if is_mark_format and not is_media_field:
            try:
                parts = val_str.split(":")
                name = parts[0].strip()
                marks_str = parts[1].strip()
                
                # Verify marks part is numeric
                try:
                    marks_val = float(marks_str)
                    
                    # Dynamic Max Marks Lookup
                    max_val = 100
                    if len(parts) >= 3:
                        try:
                            max_val = float(parts[2])
                        except:
                            max_val = subjects_config.get(name.lower().strip(), 100)
                    else:
                        max_val = subjects_config.get(name.lower().strip(), 100)
                    
                    if not max_val or max_val == 0:
                        max_val = 100
                    
                    total_obtained += marks_val
                    total_max += max_val
                    subject_marks.append({'name': name, 'marks': marks_val, 'max': max_val})
                    processed_fv_ids.add(fv.id)
                except ValueError:
                    pass
            except (IndexError):
                pass
        
        # Identify photo and signature
        elif (fv.field and fv.field.is_photo) or "photo" in label_lower or "passport" in label_lower:
            if not student_photo and fv.value != '-':
                student_photo = fv.value
            processed_fv_ids.add(fv.id)
        elif (fv.field and fv.field.is_signature) or "signature" in label_lower or "sign" in label_lower:
            if not student_signature and fv.value != '-':
                student_signature = fv.value
            processed_fv_ids.add(fv.id)

    # 3. Fetch ALL fields defined for this form to include non-filled ones
    from academics.models import FormField
    all_form_fields = FormField.objects.filter(form=application.course.form).select_related('section').order_by('section__order', 'order')
    
    # Map existing values to fields for easy lookup
    field_to_values = {}
    for fv in field_values:
        if fv.field_id not in field_to_values:
            field_to_values[fv.field_id] = []
        field_to_values[fv.field_id].append(fv)

    # Process all fields for structured display
    normal_fields = []
    for field in all_form_fields:
        values = field_to_values.get(field.id, [])
        
        if not values:
            # Create a mock fv for empty fields
            mock_fv = type('MockFV', (), {
                'field': field,
                'field_label': field.label,
                'value': '-',
                'display_value': '-',
                'id': None
            })
            values = [mock_fv]

        for fv in values:
            if fv.id and fv.id in processed_fv_ids:
                continue

            # Skip internal subject-mark fields from normal listing
            label_lower = (fv.field.label if fv.field else fv.field_label or "").lower()
            if ":" in str(fv.value) and ("mark" in label_lower or "subject" in label_lower):
                continue

            # Standard processing for display_value
            if not hasattr(fv, 'display_value'):
                val = str(fv.value).strip()
                if fv.field and fv.field.field_type in ['select', 'radio'] and val != '-':
                    from academics.models import FieldOption
                    opt = FieldOption.objects.filter(field=fv.field, value=val).first()
                    fv.display_value = opt.display_text if opt else val
                else:
                    fv.display_value = val
            
            # Special case for exam name to show the friendly name
            if ("exam" in label_lower or "qualifying" in label_lower) and "marks" not in label_lower:
                if exam_obj:
                    fv.display_value = exam_obj.name
            
            # Robust ID-to-Name resolution for Full Name fallback
            if ("name" in label_lower or "candidate" in label_lower) and (not val or ":" in val or val == "None" or not val.strip() or val == "-"):
                if val == "-":
                    fv.display_value = application.student.first_name if application.student.first_name else application.student.username
                else:
                    fv.display_value = val
            
            normal_fields.append(fv)

    percentage = (total_obtained / total_max * 100) if total_max > 0 else 0

    from applications.models import Payment
    payment = Payment.objects.filter(application=application).first()

    context = {
        'application': application,
        'app': application, 
        'field_values': normal_fields, 
        'subject_marks': subject_marks,
        'total_obtained': total_obtained,
        'total_max': total_max,
        'percentage': f"{percentage:.2f}",
        'photo': student_photo,
        'signature': student_signature,
        'payment': payment,
        'MEDIA_URL': settings.MEDIA_URL,
        'print_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
    }

    return render(request, 'institute/view_application.html', context)



# # =========================
# def view_application(request, app_id):
#     application = get_object_or_404(Application, id=app_id)
#     institute = application.course.institute if application.course else None
    
#     # Efficiently fetch field values
#     field_values = application.field_values.select_related('field', 'field__section').order_by('field__section__order', 'field__order')
    
#     student_photo = None
#     student_signature = None
#     normal_fields = []
#     subject_marks = []
    
#     for fv in field_values:
#         label_lower = fv.field.label.lower()
        
#         # Check for photo and signature specifically using field flags or label keywords
#         if fv.field.is_photo or "photo" in label_lower or "passport" in label_lower:
#             student_photo = fv.value
#         if fv.field.is_signature or "signature" in label_lower:
#             student_signature = fv.value
            
#         # Subject marks handling: Support both "Subj:Marks" and "Subj:Marks:Max"
#         if fv.value and ":" in str(fv.value) and not fv.field.is_photo and not fv.field.is_signature:
#             parts = str(fv.value).split(":")
#             if len(parts) >= 2:
#                 subject_marks.append({
#                     'subject': parts[0].strip(),
#                     'marks': parts[1].strip(),
#                     'max': parts[2].strip() if len(parts) > 2 else "100"
#                 })
#         else:
#             normal_fields.append(fv)

#     return render(request, 'institute/view_application.html', {
#         'app': application,
#         'institute': institute,
#         'field_values': normal_fields, 
#         'subject_marks': subject_marks,
#         'photo': student_photo,
#         'signature': student_signature,
#         'MEDIA_URL': settings.get('MEDIA_URL', '/media/') if hasattr(settings, 'get') else getattr(settings, 'MEDIA_URL', '/media/'),
#         'print_date': datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
#     })

# =========================
# RANKLIST 
# =========================

def calculate_total_and_percentage(application):
    total = 0
    max_total = 0
    qualified_total = 0
    main_subject_marks = 0
    sub_subject_marks = 0

    # Robust Exam Identification
    from academics.models import QualifyingExam
    exam_id = None
    for v in application.field_values.all():
        label = (v.field.label if v.field else v.field_label or "").lower()
        if ("exam" in label or "qualifying" in label) and "marks" not in label:
            val = str(v.value).strip()
            if val.isdigit():
                exam_id = int(val)
            elif val.lower().startswith('id:') and val[3:].strip().isdigit():
                exam_id = int(val[3:].strip())
            else:
                if v.field and v.field.field_type in ['select', 'radio']:
                    from academics.models import FieldOption
                    opt = FieldOption.objects.filter(field=v.field, value=val).first()
                    if opt:
                        val = opt.display_text
                
                ex_obj = QualifyingExam.objects.filter(name__iexact=val).first()
                if ex_obj:
                    exam_id = ex_obj.id
            if exam_id:
                break

    subjects_config = {}
    if exam_id:
        subjects = ExamSubject.objects.filter(exam_id=exam_id)
        for s in subjects:
            subjects_config[s.name.lower()] = {
                "include": s.include_in_rank,
                "main": s.is_main_subject,
                "sub": s.is_sub_subject,
                "max": s.max_marks,
                "pass": s.pass_mark
            }

    # calculate
    for fv in application.field_values.all():
        val_str = str(fv.value or "").strip()
        label_lower = (fv.field.label if fv.field else fv.field_label or "").lower()
        section_lower = (fv.field.section.name if fv.field and fv.field.section else fv.field_label or "").lower()
        is_qual_field = any(x in label_lower or x in section_lower for x in ["mark", "subject", "qualify", "exam"])
        if ":" in val_str and is_qual_field:
            try:
                parts = val_str.split(":")
                subject = parts[0].lower().strip()
                mark_val = float(parts[1].strip())
                
                # Dynamic Max Marks Lookup - PRIORITIZE STORED VALUE
                max_val = 100
                if len(parts) >= 3:
                    try:
                        max_val = float(parts[2])
                    except: pass
                
                if not max_val or max_val == 100:
                    config = subjects_config.get(subject, {"max": 100, "pass": 35, "include": False, "main": False, "sub": False})
                    if max_val == 100 or not max_val:
                        max_val = config.get("max", 100)
                else:
                    config = subjects_config.get(subject, {"max": 100, "pass": 35, "include": False, "main": False, "sub": False})
                
                if not max_val: max_val = 100
                
                if config.get("include", False):
                    total += mark_val
                    max_total += max_val
                    qualified_total += config.get("pass", 35)

                    if config.get("main"):
                        main_subject_marks = max(main_subject_marks, mark_val)
                    if config.get("sub"):
                        sub_subject_marks = max(sub_subject_marks, mark_val)

            except (ValueError, TypeError, IndexError):
                continue

    percentage = (total / max_total * 100) if max_total > 0 else 0
    return total, round(percentage, 2), main_subject_marks, sub_subject_marks, max_total, qualified_total

def rank_list_view(request):

    institute = request.user.institute
    course_id = request.GET.get('course')
    year_id = request.GET.get('year')
    quota_id = request.GET.get('quota')

    # Show only Verified students (status='selected')
    from django.db import models
    applications = Application.objects.filter(
        institute=institute, 
        status='selected', 
        payment__status='success'
    ).select_related('course', 'student').prefetch_related(
        models.Prefetch(
            'field_values',
            queryset=ApplicationFieldValue.objects.select_related('field', 'field__section')
        )
    )

    if course_id:
        applications = applications.filter(course_id=course_id)

    if year_id:
        applications = applications.filter(academic_year_id=year_id)

    ranked_list = []
    available_quotas = set()

    for app in applications:
        total, percentage, main_mark, sub_mark, max_total, qualified_total = calculate_total_and_percentage(app)
        quota = get_student_quota(app)

        if quota and quota != "-":
            available_quotas.add(quota)

        if quota_id and quota.lower() != quota_id.lower():
            continue

        ranked_list.append({
            "app": app,
            "name": get_student_name(app),
            "course": app.course.name if app.course else "No Course",
            "quota": quota,
            "total": total,
            "max_total": max_total,
            "qualified_total": qualified_total,
            "percentage": percentage,
            "main_mark": main_mark,
            "sub_mark": sub_mark
        })

    #  SORT DESCENDING
    ranked_list.sort(
        key=lambda x: (x['percentage'], x['main_mark'], x['sub_mark']),
        reverse=True
    )

    #  ASSIGN RANK
    for i, item in enumerate(ranked_list, start=1):
        item['rank'] = i

    # Fetch additional distinct quotas from database for complete dropdown options
    distinct_db_quotas = ApplicationFieldValue.objects.filter(
        application__institute=institute,
        field_label__icontains='quota'
    ).exclude(
        value__in=[None, '', 'None', '-']
    ).exclude(
        field_type='file'
    ).exclude(
        field__field_type='file'
    ).exclude(
        value__icontains='.pdf'
    ).exclude(
        value__icontains='.jpg'
    ).exclude(
        field_label__icontains='document'
    ).exclude(
        field_label__icontains='proof'
    ).exclude(
        field_label__icontains='certificate'
    ).exclude(
        field_label__icontains='upload'
    ).values_list('value', flat=True).distinct()

    for q in distinct_db_quotas:
        if q and q.strip() and not any(ext in q.lower() for ext in ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx']):
            available_quotas.add(q.strip())

    quotas = sorted(list(available_quotas))

    context = {
        "ranked_list": ranked_list,
        "courses": Course.objects.filter(institute=institute),
        "years": AcademicYear.objects.filter(institute=institute, is_active=True),
        "quotas": quotas,
        "selected_quota": quota_id or "",
    }

    return render(request, "institute/rank_list.html", context)


# =========================
# EXPORT RANKLIST 
# =========================
def get_student_name(application):
    """
    Returns the student's name from the User model first_name, 
    falling back to dynamic field values, and finally username.
    """
    # 1. Check User model
    if application.student and application.student.first_name:
        return application.student.first_name

    # 2. Check for field marked as is_name_field
    for v in application.field_values.all():
        if v.field and getattr(v.field, 'is_name_field', False):
            return v.value
        
    # 3. Fallback to label search
    for v in application.field_values.all():
        field_label = v.field.label if v.field else v.field_label
        if field_label:
            label = field_label.lower()
            if "name" in label and ":" not in str(v.value):
                return v.value

    if application.student:
        return application.student.username
    return f"Form #{application.id}"


def get_student_quota(application):
    """
    Extracts the quota value for an application, ignoring file uploads (e.g. death.pdf).
    """
    quota = "-"
    exact_quota_found = False

    for v in application.field_values.all():
        field_label = v.field.label if v.field else v.field_label
        if not field_label:
            continue

        lbl = field_label.lower().strip()
        val = str(v.value or "").strip()
        
        if not val or val in ["None", "-", ""]:
            continue

        if "admission quota" in lbl:
            quota = val
            exact_quota_found = True
            continue

        if "quota" in lbl and not exact_quota_found:
            # Skip document / upload field labels
            if any(word in lbl for word in ['document', 'proof', 'certificate', 'upload', 'file', 'pdf', 'image']):
                continue
                
            # Exclude file uploads
            is_file = False
            if hasattr(v, 'field_type') and v.field_type == 'file':
                is_file = True
            if v.field and hasattr(v.field, 'field_type') and v.field.field_type == 'file':
                is_file = True
            
            val_lower = val.lower()
            if any(ext in val_lower for ext in ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx']):
                is_file = True
                
            if not is_file:
                quota = val

    # Fallback to admission object if present
    if quota == "-" and hasattr(application, 'admission') and application.admission and getattr(application.admission, 'admission_quota', None):
        quota = application.admission.admission_quota

    return quota


@login_required
def export_rank_excel(request):
    institute = getattr(request.user, 'institute', None)
    if not institute:
        return HttpResponse("Unauthorized", status=401)
    course_id = request.GET.get('course')
    year_id = request.GET.get('year')
    quota_id = request.GET.get('quota')

    # Show only Verified students (status='selected')
    from django.db import models
    applications = Application.objects.filter(
        institute=institute, 
        status='selected', 
        payment__status='success'
    ).select_related('course', 'student').prefetch_related(
        models.Prefetch(
            'field_values',
            queryset=ApplicationFieldValue.objects.select_related('field', 'field__section')
        )
    )

    if course_id:
        applications = applications.filter(course_id=course_id)

    if year_id:
        applications = applications.filter(academic_year_id=year_id)

    ranked_list = []
    unique_subjects = []

    for app in applications:
        try:
            total, percentage, main_mark, sub_mark, max_total, qualified_total = calculate_total_and_percentage(app)
            
            # Extract student details and subjects
            mobile = "-"
            gender = "-"
            exam_name = "-"
            quota = get_student_quota(app)

            if quota_id and quota.lower() != quota_id.lower():
                continue

            choice1 = "-"
            choice2 = "-"
            choice3 = "-"
            subjects_data = {} # {name: {marks, max, pass}}
            
            # We need subjects_config for pass marks
            # First find exam_id
            exam_id = None
            for v in app.field_values.all():
                label = (v.field.label if v.field else v.field_label or "").lower()
                if ("exam" in label or "qualifying" in label) and "marks" not in label:
                    val = str(v.value).strip()
                    if val.isdigit(): exam_id = int(val)
                    elif val.lower().startswith('id:') and val[3:].strip().isdigit(): exam_id = int(val[3:].strip())
                    else:
                        from academics.models import QualifyingExam
                        ex_obj = QualifyingExam.objects.filter(name__iexact=val).first()
                        if ex_obj: exam_id = ex_obj.id
                    if exam_id: break

            subjects_config = {}
            if exam_id:
                from academics.models import ExamSubject
                for s in ExamSubject.objects.filter(exam_id=exam_id):
                    subjects_config[(s.name or "").lower().strip()] = {
                        "max": s.max_marks,
                        "pass": s.pass_mark,
                        "full_name": s.name
                    }

            for v in app.field_values.all():
                lbl = (v.field.label if v.field else v.field_label or "").lower()
                val = str(v.value or "").strip()
                
                if "mobile" in lbl or "phone" in lbl or "contact" in lbl:
                    mobile = val
                elif "gender" in lbl:
                    gender = val
                elif "choice 1" in lbl or "choice1" in lbl:
                    choice1 = val
                elif "choice 2" in lbl or "choice2" in lbl:
                    choice2 = val
                elif "choice 3" in lbl or "choice3" in lbl:
                    choice3 = val
                elif ("exam" in lbl or "qualifying" in lbl) and "marks" not in lbl:
                    if exam_id:
                        from academics.models import QualifyingExam
                        ex = QualifyingExam.objects.filter(id=exam_id).first()
                        exam_name = ex.name if ex else val
                    else:
                        exam_name = val
                
                if ":" in val:
                    section_lower = (v.field.section.name if v.field and v.field.section else v.field_label or "").lower()
                    is_qual_field = any(x in lbl or x in section_lower for x in ["mark", "subject", "qualify", "exam"])
                    if is_qual_field:
                        try:
                            parts = val.split(":")
                            if len(parts) >= 2:
                                s_name_raw = parts[0].strip()
                                s_marks = parts[1].strip()
                                
                                config = subjects_config.get(s_name_raw.lower(), {"max": 100, "pass": 35, "full_name": s_name_raw})
                                s_max = parts[2].strip() if len(parts) > 2 else str(config["max"])
                                s_pass = str(config["pass"])
                                
                                s_display_name = config["full_name"]
                                subjects_data[s_display_name] = {
                                    "marks": s_marks,
                                    "max": s_max,
                                    "pass": s_pass
                                }
                                if s_display_name not in unique_subjects:
                                    unique_subjects.append(s_display_name)
                        except: pass

            ranked_list.append({
                "name": get_student_name(app),
                "mobile": mobile,
                "gender": gender,
                "exam_name": exam_name,
                "subjects": subjects_data,
                "total": total,
                "max_total": max_total,
                "percentage": percentage,
                "main_mark": main_mark,
                "sub_mark": sub_mark,
                # Extra fields
                "date": app.created_at.strftime('%Y-%m-%d') if app.created_at else "-",
                "form_id": app.id,
                "quota": quota,
                "choice1": choice1,
                "choice2": choice2,
                "choice3": choice3
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error exporting application {app.id} to excel: {str(e)}", exc_info=True)
            continue

    # SORT DESCENDING
    ranked_list.sort(
        key=lambda x: (x['percentage'], x['main_mark'], x['sub_mark']),
        reverse=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rank List"

    # HEADER
    header = ["Rank", "Name", "Mobile", "Gender", "Qualifying Examination"]
    # Subject Headers
    for sub in unique_subjects:
        header.extend([f"{sub} (Obtained)", f"{sub} (Out Of)", f"{sub} (Min Mark)"])
    header.extend(["Total Marks Obtained", "Maximum Marks Total", "Percentage", "Date", "Form No", "Student Name", "Mobile No.", "Quota", "Gender", "Choice1", "Choice2", "Choice3"])
    
    ws.append(header)

    # DATA
    for i, item in enumerate(ranked_list, start=1):
        row = [
            i,
            item['name'],
            item['mobile'],
            item['gender'],
            item['exam_name']
        ]
        for sub in unique_subjects:
            s_info = item['subjects'].get(sub, {"marks": "-", "max": "-", "pass": "-"})
            row.extend([s_info['marks'], s_info['max'], s_info['pass']])
            
        row.extend([
            item['total'],
            item['max_total'],
            f"{item['percentage']}%",
            item['date'],
            item['form_id'],
            item['name'],
            item['mobile'],
            item['quota'],
            item['gender'],
            item['choice1'],
            item['choice2'],
            item['choice3']
        ])
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=rank_list.xlsx'

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.content = output.read()

    return response

# =========================
# EXPORT STUDENT LIST
# =========================
@login_required
def excel_export_students(request):
    institute = getattr(request.user, 'institute', None)
    if not institute:
        return HttpResponse("Unauthorized", status=401)

    query = request.GET.get('q', '')
    course_filter = request.GET.get('course', '')
    year_filter = request.GET.get('year', '')
    status_filter = request.GET.get('status', '')

    admissions = Admission.objects.filter(application__institute=institute).select_related(
        'application', 'selected_course', 'application__academic_year', 'joining_period'
    ).prefetch_related(
        'application__field_values', 'application__field_values__field'
    ).order_by('-created_at')

    if query:
        admissions = admissions.filter(
            Q(register_number__icontains=query) |
            Q(application__id__icontains=query) |
            Q(application__field_values__value__icontains=query, application__field_values__field__is_name_field=True)
        ).distinct()

    if course_filter:
        admissions = admissions.filter(selected_course_id=course_filter)
    if year_filter:
        admissions = admissions.filter(application__academic_year_id=year_filter)
    if status_filter:
        admissions = admissions.filter(status=status_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Inventory"

    headers = ["SL NO", "Register No", "Student Name", "Course", "Year/Period", "Admission Year", "Status"]
    ws.append(headers)

    for i, adm in enumerate(admissions, start=1):
        ws.append([
            i,
            adm.register_number,
            get_student_name(adm.application),
            adm.selected_course.name,
            adm.joining_period.name if adm.joining_period else "N/A",
            adm.application.academic_year.name if adm.application.academic_year else "N/A",
            adm.get_status_display()
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=Students_Export_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

# =========================
# REGISTER
# =========================
def institute_register(request):
    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')
        institute_name = request.POST.get('institute_name')

        user = User.objects.create_user(
            username=username,
            password=password,
            role='institute'
        )

        institute = Institute.objects.create(
            user=user,
            name=institute_name,
            code=username.upper()
        )

        # link reverse
        user.institute = institute
        user.save()

        return redirect('/institute/login/')

    return render(request, 'institute/register.html')


# =========================
# DASHBOARD
# =========================
@login_required
def institute_dashboard(request):
    institute = getattr(request.user, 'institute', None)
    if not institute:
        if request.user.is_staff or request.user.is_superuser:
            institute = Institute.objects.first()
        else:
            return redirect('/institute/register/')
    if not institute:
        return redirect('/')

    # Support filtering and search
    query = request.GET.get('q', '')
    course_filter = request.GET.get('course', '')
    year_filter = request.GET.get('year', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    fee_category_filter = request.GET.get('fee_category', '')
    quota_filter = request.GET.get('quota', '')

    # REQUIREMENT: Only students with Payment Status = Paid should appear
    apps = Application.objects.filter(
        institute=institute, 
        payment__status='success'
    ).select_related(
        'course', 'academic_year', 'admission'
    ).prefetch_related(
        'field_values', 'field_values__field'
    ).order_by('-created_at')

    if query:
        apps = apps.filter(
            Q(id__icontains=query) |
            Q(admission__register_number__icontains=query) |
            Q(field_values__value__icontains=query, field_values__field__is_name_field=True)
        ).distinct()

    if course_filter:
        apps = apps.filter(course_id=course_filter)
    if year_filter:
        apps = apps.filter(academic_year_id=year_filter)
    
    # Status filter is combined (Application status vs Admission status)
    if status_filter:
        if status_filter in ['active', 'warned', 'suspended', 'trashed']:
            apps = apps.filter(admission__status=status_filter)
        elif status_filter in ['submitted', 'pending', 'selected', 'rejected', 'hold']:
            apps = apps.filter(status=status_filter)
        else:
            # Try both if unknown
            apps = apps.filter(Q(status=status_filter) | Q(admission__status=status_filter))

    if date_from:
        apps = apps.filter(created_at__date__gte=date_from)
    if date_to:
        apps = apps.filter(created_at__date__lte=date_to)
    if fee_category_filter:
        apps = apps.filter(selected_fee_type__name=fee_category_filter)
    if quota_filter:
        apps = apps.filter(
            field_values__field_label__icontains='quota',
            field_values__value=quota_filter
        )

    # Performance: Pagination
    paginator = Paginator(apps, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Process names and fields for display
    processed_admissions = []
    for app in page_obj:
        name = get_student_name(app)
        
        # Extract Contact, Caste, Gender, etc.
        contact = "-"
        gender = "-"
        quota = "-"
        remarks = app.remarks or ""
        
        exact_quota_found = False

        # Optimized field extraction
        for v in app.field_values.all():
            field_label = v.field.label if v.field else v.field_label
            if not field_label:
                continue

            lbl = field_label.lower().strip()
            val = v.value
            
            if not val or str(val).strip() == "None": continue

            if "admission quota" in lbl:
                quota = str(val).strip()
                exact_quota_found = True
                continue

            if "phone" in lbl or "mobile" in lbl or "contact" in lbl:
                contact = val
            elif "gender" in lbl:
                gender = val
            elif "quota" in lbl and not exact_quota_found:
                # Explicitly skip if the label asks for a document, proof, etc.
                if any(word in lbl for word in ['document', 'proof', 'certificate', 'upload', 'file', 'pdf', 'image']):
                    continue
                    
                # Exclude file uploads from being captured as the 'quota' string
                is_file = False
                if hasattr(v, 'field_type') and v.field_type == 'file':
                    is_file = True
                if v.field and hasattr(v.field, 'field_type') and v.field.field_type == 'file':
                    is_file = True
                
                val_lower = str(val).strip().lower()
                if any(ext in val_lower for ext in ['.pdf', '.jpg', '.jpeg', '.png', '.doc']):
                    is_file = True
                    
                if not is_file:
                    quota = val
            elif "remarks" in lbl or "comment" in lbl:
                if not remarks:
                    remarks = val

        processed_admissions.append({
            'form_id': app.id,
            'student_name': name,
            'contact': contact,
            'gender': gender,
            'quota': quota,
            'fee_category': app.selected_fee_type.name if app.selected_fee_type else "-",
            'remarks': remarks,
            'status': app.status,
            'status_display': app.get_status_display(),
        })

    courses = Course.objects.filter(institute=institute)
    years = AcademicYear.objects.filter(institute=institute, is_active=True)
    
    from academics.models import ApplicationFeeType
    fee_categories = ApplicationFeeType.objects.filter(form__course__institute=institute, is_active=True).values_list('name', flat=True).distinct()
    if not fee_categories:
        fee_categories = ApplicationFeeType.objects.filter(is_active=True).values_list('name', flat=True).distinct()

    # Get distinct quotas for filter dropdown
    distinct_quotas = ApplicationFieldValue.objects.filter(
        application__institute=institute,
        field_label__icontains='quota'
    ).exclude(
        value__in=[None, '', 'None', '-']
    ).exclude(
        field_type='file'
    ).exclude(
        field__field_type='file'
    ).exclude(
        value__icontains='.pdf'
    ).exclude(
        value__icontains='.jpg'
    ).values_list('value', flat=True).distinct()
    quotas = sorted(list(set(q.strip() for q in distinct_quotas if q and q.strip())))

    # REQUIREMENT: Notice Board visibility for all users
    from academics.models import NoticeBoard
    notices = NoticeBoard.objects.filter(
        is_active=True,
        course__isnull=True,
        assigned_class__isnull=True
    ).order_by('-created_at')[:10]

    return render(request, 'institute/dashboard.html', {
        'page_obj': page_obj,
        'admissions': processed_admissions,
        'courses': courses,
        'years': years,
        'fee_categories': fee_categories,
        'quotas': quotas,
        'selected_course': course_filter,
        'selected_year': year_filter,
        'selected_status': status_filter,
        'selected_fee_category': fee_category_filter,
        'selected_quota': quota_filter,
        'date_from': date_from,
        'date_to': date_to,
        'query': query,
        'notices': notices
    })

# =========================
# STD NAME EXTRACT
# =========================
# Name extraction handled by the global helper function at line 842.




# =========================
# EDIT APPLICATION
# =========================
@login_required
def edit_application(request, app_id):

    app = get_object_or_404(Application, id=app_id)

    # =========================
    # GET FORM FIELDS
    # =========================
    fields = []
    if hasattr(app.course, 'form') and app.course.form:
        fields = FormField.objects.filter(
            form=app.course.form
        ).select_related('section').order_by('section__order', 'order')

    # ATTACH VALUES TO FIELDS
    # Use order_by('id') so that if duplicates exist, the latest one (highest ID) is kept in the dictionary
    # CRITICAL: Exclude snapshot values (containing :) from field_values used for form rendering
    field_values = {}
    for v in app.field_values.filter(field__isnull=False).order_by('id'):
        val_str = str(v.value or "")
        if ":" not in val_str: 
            field_values[v.field_id] = v.value

    for f in fields:
        f.current_value = field_values.get(f.id, "")
        label_lower = f.label.lower()

        # FIX: Ensure Full Name shows student name, not corrupted subject marks
        if f.label == "Full Name" and (not f.current_value or ":" in str(f.current_value)):
            f.current_value = app.student.first_name
        
        # Resolve Field Options for non-choice fields (e.g. text fields used as ID holders)
        if f.field_type not in ['select', 'radio', 'checkbox'] and f.current_value:
            from academics.models import FieldOption
            opt = FieldOption.objects.filter(field=f, value=f.current_value).first()
            if opt:
                f.current_value = opt.display_text
        
        # REQUIREMENT: Choice 3 must remain blank if not filled
        if not f.current_value or str(f.current_value).lower() in ['none', 'null', 'select', '', '-', 'empty']:
            f.current_value = ""

        f.value = f.current_value  # Support templates using .value or .current_value

    # =========================
    # SAVE (POST)
    # =========================
    if request.method == 'POST':

        for field in fields:    
            key = f'field_{field.id}'
            
            # FILE FIELD
            if field.field_type == 'file':
                file_obj = request.FILES.get(key)

                if file_obj:
                    # Clean up duplicates to avoid MultipleObjectsReturned
                    ApplicationFieldValue.objects.filter(application=app, field=field).delete()
                    from django.core.files.storage import FileSystemStorage
                    fs = FileSystemStorage()
                    filename = fs.save(file_obj.name, file_obj)
                    ApplicationFieldValue.objects.create(
                        application=app,
                        field=field,
                        field_label=field.label,
                        field_type=field.field_type,
                        value=filename
                    )

            else:
                val = request.POST.get(key)

                if val is not None:
                    # Clean up duplicates to avoid MultipleObjectsReturned
                    # and ensure we only have one record per field
                    fvs = ApplicationFieldValue.objects.filter(application=app, field=field)
                    if fvs.count() > 1:
                        fvs.delete()
                        ApplicationFieldValue.objects.create(
                            application=app,
                            field=field,
                            field_label=field.label,
                            field_type=field.field_type,
                            value=val
                        )
                    else:
                        ApplicationFieldValue.objects.update_or_create(
                            application=app,
                            field=field,
                            defaults={
                                'value': val,
                                'field_label': field.label,
                                'field_type': field.field_type
                            }
                        )

        # =========================
        #  UPDATE SUBJECTS (FIXED)
        # =========================
        # =========================
        #  UPDATE SUBJECTS (Parsing correctly)
        # =========================
        # Delete only "marks" type values to replace them
        ApplicationFieldValue.objects.filter(
            application=app,
            value__contains=":"
        ).exclude(
            field__field_type='file' # Don't delete file values which might have colons (rare but possible)
        ).delete()

        # Only target fields in the 'Qualifying Examination' section or similar
        qe_field = FormField.objects.filter(form=app.course.form, section__name__icontains="Qualifi").first()
        if not qe_field:
            qe_field = FormField.objects.filter(form=app.course.form, label__icontains="Qualifi").first()
        if not qe_field:
            qe_field = FormField.objects.filter(form=app.course.form, section__name__icontains="mark").first()

        for key in request.POST:
            if key.startswith("subject_"):
                subject_name = key.replace("subject_", "").strip()
                marks = request.POST.get(key)
                max_val = request.POST.get(f"max_{subject_name}", "100")
                if marks is not None and marks != "": # Save even if 0
                    ApplicationFieldValue.objects.create(
                        application=app,
                        field=qe_field,
                        field_label=qe_field.label if qe_field else "Qualifying Examination Marks",
                        field_type="text", # Snapshot fields
                        value=f"{subject_name}:{marks}:{max_val}"
                    )

        # =========================
        # STATUS + REMARKS
        # =========================
        old_status = app.status
        app.status = request.POST.get('status')
        app.remarks = request.POST.get('remarks')
        app.save()

        # Trigger Email if status changed
        if old_status != app.status:
            send_status_email(app, app.status)

        return redirect('/institute/dashboard/')

    # 1. Identify Qualifying Exam and Subjects Configuration
    exam_id = None
    exam_obj = None
    subjects_config = {}

    # Identify the examination from form values
    for fv in app.field_values.all().order_by('-id'): # Check latest first
        label = (fv.field.label if fv.field else fv.field_label or "").lower()
        val = str(fv.value).strip()
        
        # Skip snapshots when looking for the Exam ID
        if ":" in val: continue

        if ("exam" in label or "qualifying" in label) and "marks" not in label:
            # If it's a choice/code, resolve the display name first
            if fv.field and fv.field.field_type in ['select', 'radio']:
                from academics.models import FieldOption
                opt = FieldOption.objects.filter(field=fv.field, value=val).first()
                if opt:
                    val = opt.display_text

            if val.isdigit():
                exam_id = int(val)
            else:
                from academics.models import QualifyingExam
                ex = QualifyingExam.objects.filter(name__iexact=val).first()
                if ex: 
                    exam_id = ex.id
                    exam_obj = ex
            if exam_id: break

    if exam_id and not exam_obj:
        from academics.models import QualifyingExam
        exam_obj = QualifyingExam.objects.filter(id=exam_id).first()

    if exam_obj:
        from academics.models import ExamSubject
        for s in ExamSubject.objects.filter(exam=exam_obj):
            subjects_config[s.name.lower().strip()] = {
                "max": s.max_marks,
                "pass": s.pass_mark,
                "name": s.name
            }

    # 2. Extract ALL stored marks (exact copy requirement)
    # We use a dictionary to keep the latest value for each subject name
    stored_marks = {}
    for v in app.field_values.all():
        val_str = str(v.value or "").strip()
        # Same logic as view_application: check for colon and numeric second part
        label_lower = (v.field.label if v.field else v.field_label or "").lower()
        section_lower = (v.field.section.name if v.field and v.field.section else v.field_label or "").lower()
        is_qual_field = any(x in label_lower or x in section_lower for x in ["mark", "subject", "qualify", "exam"])
        if ":" in val_str and is_qual_field:
            label_lower = (v.field.label if v.field else v.field_label or "").lower()
            is_media = any(x in label_lower for x in ["photo", "signature", "sign"])
            if is_media: continue

            parts = val_str.split(":")
            if len(parts) >= 2:
                s_name = parts[0].strip()
                s_marks = parts[1].strip()
                s_max = parts[2].strip() if len(parts) >= 3 else None
                
                # Check if second part is numeric to avoid false positives
                try:
                    float(s_marks)
                    stored_marks[s_name.lower().strip()] = {
                        "name": s_name,
                        "marks": s_marks,
                        "max": s_max
                    }
                except ValueError:
                    continue

    # 3. Prepare final subjects list
    subjects = []
    processed_names = set()

    # Priority 1: Subjects from Exam Configuration (maintains order)
    if exam_obj:
        from academics.models import ExamSubject
        for es in ExamSubject.objects.filter(exam=exam_obj).order_by('id'):
            es_name_lower = es.name.lower().strip()
            val = stored_marks.get(es_name_lower)
            
            marks = val["marks"] if val else ""
            # Prioritize stored max marks, fallback to config
            max_val = val["max"] if val and val["max"] else es.max_marks
            
            subjects.append({
                "name": es.name,
                "marks": marks,
                "max": max_val,
                "pass": es.pass_mark
            })
            processed_names.add(es_name_lower)

    # Priority 2: Any other marks found in DB but not in config (the "exact copy" part)
    for name_lower, data in stored_marks.items():
        if name_lower not in processed_names:
            config = subjects_config.get(name_lower, {"max": 100, "pass": 35})
            subjects.append({
                "name": data["name"],
                "marks": data["marks"],
                "max": data["max"] or config["max"],
                "pass": config["pass"]
            })
        
        

    # =========================
    # RENDER
    # =========================
    return render(request, 'institute/edit_application.html', {
        'app': app,
        'fields': fields,
        'subjects': subjects,
    })

# Note: update_student_status is defined below under Student List section


# ✅ EXCEL TEMPLATE DOWNLOAD
@login_required
def download_excel_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import Template"
    
    headers = [
        "Full Name *", "Mobile Number *", "Email Address *", "Registration ID *", 
        "Date of Join *", "Admission Quota *", "Academic Session *", "Select Course *",
        "Fee Category", "Assign Student to Class", "Assign Student to Class Year / Semester",
        "Joining Period (Excludes Previous Fees)",
        "Care Of", "Guardian Name", "Guardian Mobile", "Relationship", "Guardian Address"
    ]
    
    # Add dynamic form fields as extra columns if available
    dynamic_fields = FormField.objects.filter(form__course__institute=request.user.institute).values_list('label', flat=True).distinct()
    for field_label in dynamic_fields:
        if field_label not in headers:
            headers.append(f"Field: {field_label}")

    ws.append(headers)
    
    # Add sample row matching requested fields
    sample_row = [
        "ABC", "9856565656", "abc@gmail.com", "123",
        "01-06-2026", "Management", "2026-27", "Bachelor of Pharmacy",
        "GENERAL", "B.Pharm", "1 SEMESTER", "1st Year",
        "", "", "", "", ""
    ]
    ws.append(sample_row)

    # Adjust column widths
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    return response


# ✅ EXCEL BULK IMPORT
@login_required
def excel_import_students(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return JsonResponse({'status': 'error', 'message': 'The Excel file contains no data rows.'}, status=400)

            raw_headers = rows[0]
            header_map = {}
            field_cols = {}

            for idx, h in enumerate(raw_headers):
                if not h:
                    continue
                h_str = str(h).strip()
                h_clean = h_str.lower().replace('*', '').replace('(', '').replace(')', '').strip()
                
                if h_str.startswith("Field: "):
                    field_cols[h_str.replace("Field: ", "").strip()] = idx
                elif 'full name' in h_clean or 'student name' in h_clean:
                    header_map['full_name'] = idx
                elif 'mobile' in h_clean or 'phone' in h_clean:
                    header_map['mobile'] = idx
                elif 'email' in h_clean:
                    header_map['email'] = idx
                elif 'registration' in h_clean or 'reg id' in h_clean or 'reg_id' in h_clean:
                    header_map['registration_id'] = idx
                elif 'date of join' in h_clean or 'join' in h_clean or 'doj' in h_clean:
                    header_map['date_of_join'] = idx
                elif 'quota' in h_clean:
                    header_map['admission_quota'] = idx
                elif 'academic session' in h_clean or 'academic year' in h_clean or 'session' in h_clean:
                    header_map['academic_session'] = idx
                elif 'select course' in h_clean or 'course' in h_clean:
                    header_map['course'] = idx
                elif 'fee category' in h_clean:
                    header_map['fee_category'] = idx
                elif 'joining period' in h_clean or 'period' in h_clean:
                    header_map['joining_period'] = idx
                elif 'class year' in h_clean or 'semester' in h_clean:
                    header_map['class_year'] = idx
                elif 'class' in h_clean:
                    header_map['class'] = idx
                elif 'care of' in h_clean:
                    header_map['care_of'] = idx
                elif 'guardian name' in h_clean:
                    header_map['guardian_name'] = idx
                elif 'guardian mobile' in h_clean:
                    header_map['guardian_mobile'] = idx
                elif 'relationship' in h_clean:
                    header_map['relationship'] = idx
                elif 'guardian address' in h_clean or 'address' in h_clean:
                    header_map['guardian_address'] = idx

            data_rows = rows[1:]
            report = {'success': 0, 'errors': []}
            institute = request.user.institute
            seen_reg_ids_in_batch = set()

            def get_val(row_tuple, key, default_idx=None):
                if key in header_map and header_map[key] < len(row_tuple):
                    v = row_tuple[header_map[key]]
                    if v is not None and str(v).strip() != "":
                        return v
                if default_idx is not None and default_idx < len(row_tuple):
                    v = row_tuple[default_idx]
                    if v is not None and str(v).strip() != "":
                        return v
                return None

            for row_idx, row in enumerate(data_rows, start=2):
                if not any(row):
                    continue # Skip blank rows

                try:
                    full_name = str(get_val(row, 'full_name', 0) or '').strip()
                    mobile_raw = get_val(row, 'mobile', 1)
                    if isinstance(mobile_raw, (float, int)):
                        mobile = str(int(mobile_raw)).strip()
                    else:
                        mobile = str(mobile_raw or '').strip()

                    email = str(get_val(row, 'email', 2) or '').strip()
                    reg_id = str(get_val(row, 'registration_id', 3) or '').strip()
                    doj_raw = get_val(row, 'date_of_join', 4)
                    quota_raw = str(get_val(row, 'admission_quota', 5) or '').strip()
                    session_raw = str(get_val(row, 'academic_session', 6) or '').strip()
                    course_raw = str(get_val(row, 'course', 7) or '').strip()
                    fee_cat_raw = str(get_val(row, 'fee_category', 8) or '').strip()
                    class_raw = str(get_val(row, 'class', 9) or '').strip()
                    class_year_raw = str(get_val(row, 'class_year', 10) or '').strip()
                    joining_period_raw = str(get_val(row, 'joining_period', 11) or '').strip()

                    care_of = str(get_val(row, 'care_of', 12) or '').strip()
                    g_name = str(get_val(row, 'guardian_name', 13) or '').strip()
                    g_mobile_raw = get_val(row, 'guardian_mobile', 14)
                    if isinstance(g_mobile_raw, (float, int)):
                        g_mobile = str(int(g_mobile_raw)).strip()
                    else:
                        g_mobile = str(g_mobile_raw or '').strip()
                    rel = str(get_val(row, 'relationship', 15) or '').strip()
                    g_addr = str(get_val(row, 'guardian_address', 16) or '').strip()

                    if not full_name:
                        report['errors'].append(f"Row {row_idx}: Missing Full Name.")
                        continue

                    if not mobile:
                        mobile = f"STU{row_idx}{datetime.datetime.now().strftime('%M%S')}"

                    if not reg_id:
                        reg_id = f"REG_{mobile}"

                    # Strict Registration ID duplicate validation
                    if Admission.objects.filter(registration_id=reg_id).exists() or reg_id in seen_reg_ids_in_batch:
                        report['errors'].append(f"Row {row_idx}: Duplicate Registration ID '{reg_id}'. This Registration ID already exists in the system or is duplicated in the file.")
                        continue

                    seen_reg_ids_in_batch.add(reg_id)

                    # Course lookup with flexible fallback
                    course = None
                    if course_raw:
                        course = Course.objects.filter(institute=institute).filter(
                            Q(name__iexact=course_raw) | Q(course_code__iexact=course_raw)
                        ).first()
                        if not course:
                            course = Course.objects.filter(institute=institute).filter(
                                Q(name__icontains=course_raw) | Q(course_code__icontains=course_raw)
                            ).first()
                        if not course:
                            cls_match = Class.objects.filter(institute=institute, name__icontains=course_raw).first()
                            if cls_match and cls_match.course:
                                course = cls_match.course

                    if not course:
                        course = Course.objects.filter(institute=institute).first()

                    if not course:
                        report['errors'].append(f"Row {row_idx}: No course found for institute.")
                        continue

                    # Academic Year lookup
                    year_obj = None
                    if session_raw:
                        year_obj = AcademicYear.objects.filter(institute=institute).filter(
                            Q(name__iexact=session_raw) | Q(id__iexact=session_raw) | Q(name__icontains=session_raw)
                        ).first()
                    if not year_obj:
                        year_obj = AcademicYear.objects.filter(institute=institute, is_active=True).first()

                    # Date of Join parsing
                    if isinstance(doj_raw, (datetime.date, datetime.datetime)):
                        doj_obj = doj_raw.date() if isinstance(doj_raw, datetime.datetime) else doj_raw
                    elif isinstance(doj_raw, (int, float)):
                        try:
                            from openpyxl.utils.datetime import from_excel
                            doj_obj = from_excel(doj_raw).date()
                        except Exception:
                            doj_obj = datetime.date.today()
                    elif doj_raw:
                        doj_str = str(doj_raw).strip()
                        parsed_date = None
                        for fmt in ['%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y']:
                            try:
                                parsed_date = datetime.datetime.strptime(doj_str, fmt).date()
                                break
                            except ValueError:
                                pass
                        doj_obj = parsed_date or datetime.date.today()
                    else:
                        doj_obj = datetime.date.today()

                    # Quota
                    admission_quota = quota_raw if quota_raw in ['Merit', 'Management', 'NRI'] else 'Merit'

                    # Fee Category lookup
                    fee_cat_master = None
                    app_fee_cat = None
                    app_fee_type = None
                    if fee_cat_raw:
                        fee_cat_master = FeeCategoryMaster.objects.filter(is_active=True).filter(
                            Q(name__iexact=fee_cat_raw) | Q(name__icontains=fee_cat_raw)
                        ).first()
                        app_fee_cat = FeeCategory.objects.filter(course=course).filter(
                            Q(name__iexact=fee_cat_raw) | Q(name__icontains=fee_cat_raw)
                        ).first()
                        app_fee_type = ApplicationFeeType.objects.filter(form__course=course).filter(
                            Q(name__iexact=fee_cat_raw) | Q(name__icontains=fee_cat_raw)
                        ).first()

                    if not app_fee_cat and course:
                        app_fee_cat = FeeCategory.objects.filter(course=course).first()

                    # Joining Period lookup
                    joining_period_obj = None
                    if joining_period_raw:
                        jp_qs = CourseSubCategory.objects.filter(
                            Q(name__iexact=joining_period_raw) | Q(name__icontains=joining_period_raw)
                        )
                        if course and course.category:
                            jp_qs = jp_qs.filter(category=course.category)
                        joining_period_obj = jp_qs.first()

                    # Class lookup
                    assigned_class = None
                    if class_raw:
                        assigned_class = Class.objects.filter(institute=institute).filter(
                            Q(name__iexact=class_raw) | Q(name__icontains=class_raw)
                        ).first()
                    if not assigned_class and course:
                        assigned_class = Class.objects.filter(institute=institute, course=course).first()

                    # Class Year / Semester lookup
                    assigned_class_year = None
                    if class_year_raw:
                        cy_qs = ClassYear.objects.filter(is_active=True)
                        if assigned_class:
                            cy_qs = cy_qs.filter(class_obj=assigned_class)
                        assigned_class_year = cy_qs.filter(
                            Q(name__iexact=class_year_raw) | Q(name__icontains=class_year_raw)
                        ).first()

                    # Create/get User
                    user, created = User.objects.get_or_create(
                        username=mobile,
                        defaults={'email': email or f"{mobile}@jdt.local", 'role': 'student', 'first_name': full_name}
                    )
                    if not created:
                        if full_name and user.first_name != full_name:
                            user.first_name = full_name
                        if email and user.email != email:
                            user.email = email
                        user.role = 'student'
                        user.save()

                    # Create Application
                    app = Application.objects.create(
                        student=user,
                        institute=institute,
                        academic_year=year_obj,
                        course=course,
                        selected_fee_type=app_fee_type,
                        status='selected'
                    )

                    # Calculate total fee
                    total_fee_val = app_fee_cat.total_fee if app_fee_cat else (app_fee_type.amount if app_fee_type else 0.00)

                    # Create Payment record with status 'success'
                    Payment.objects.get_or_create(
                        application=app,
                        defaults={
                            'amount': total_fee_val,
                            'status': 'success',
                            'payment_mode': 'CASH',
                            'payment_date': doj_obj
                        }
                    )

                    # Snapshot standard & dynamic field values for application audit & searching
                    standard_field_snapshots = [
                        ("Full Name", full_name),
                        ("Mobile Number", mobile),
                        ("Email Address", email),
                        ("Registration ID", reg_id),
                        ("Date of Join", str(doj_obj)),
                        ("Admission Quota", admission_quota),
                        ("Academic Session", session_raw or (year_obj.name if year_obj else '')),
                        ("Course", course.name if course else course_raw),
                        ("Fee Category", fee_cat_raw),
                        ("Assign Student to Class", assigned_class.name if assigned_class else class_raw),
                        ("Assign Student to Class Year / Semester", assigned_class_year.name if assigned_class_year else class_year_raw),
                        ("Joining Period", joining_period_obj.name if joining_period_obj else joining_period_raw),
                    ]

                    for label, val in standard_field_snapshots:
                        if val:
                            matching_form_field = FormField.objects.filter(form__course=course, label__iexact=label).first()
                            ApplicationFieldValue.objects.create(
                                application=app,
                                field=matching_form_field,
                                field_label=matching_form_field.label if matching_form_field else label,
                                field_type=matching_form_field.field_type if matching_form_field else 'text',
                                value=str(val)
                            )

                    # Save Extra Dynamic Fields if present
                    form_fields = FormField.objects.filter(form__course=course)
                    standard_labels = [s[0].lower() for s in standard_field_snapshots]
                    for f in form_fields:
                        if f.label in field_cols and f.label.lower() not in standard_labels:
                            val = row[field_cols[f.label]]
                            if val is not None:
                                ApplicationFieldValue.objects.create(
                                    application=app,
                                    field=f,
                                    field_label=f.label,
                                    field_type=f.field_type,
                                    value=str(val)
                                )

                    # Create Admission Record
                    Admission.objects.create(
                        application=app,
                        registration_id=reg_id,
                        admission_quota=admission_quota,
                        date_of_join=doj_obj,
                        selected_course=course,
                        joining_period=joining_period_obj,
                        fee_category=app_fee_cat,
                        assigned_fee_category=fee_cat_master,
                        assigned_class=assigned_class,
                        assigned_class_year=assigned_class_year,
                        calculated_fee=total_fee_val,
                        discount_amount=0.00,
                        final_fee=total_fee_val,
                        care_of=care_of,
                        guardian_name=g_name or full_name,
                        guardian_mobile=g_mobile or mobile,
                        relationship=rel or '',
                        guardian_address=g_addr or ''
                    )

                    report['success'] += 1

                except Exception as e:
                    report['errors'].append(f"Row {row_idx}: {str(e)}")

            return JsonResponse({'status': 'success', 'report': report})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"Failed to process file: {str(e)}"}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


# ✅ LOGOUT (COMMON FOR ALL USERS )
def user_logout(request):
    logout(request)
    return redirect('institute_login')

# =========================
# STUDENT LIST & EXPORT
# =========================


@login_required
def student_list_view(request):
    institute = request.user.institute
    admissions = Admission.objects.filter(application__institute=institute).select_related('application__student', 'application__academic_year', 'application__course').prefetch_related('uploaded_documents__uploaded_by')

    # Filters
    form_id = request.GET.get('form_id') # Search by Admission No / Register Number
    name = request.GET.get('name')
    batch_id = request.GET.get('batch_id')
    course_id = request.GET.get('course_id')
    status_filter = request.GET.get('status')
    class_id = request.GET.get('class_id')
    class_year_id = request.GET.get('class_year_id')

    if form_id:
        admissions = admissions.filter(Q(register_number__icontains=form_id) | Q(registration_id__icontains=form_id))
    
    if name:
        admissions = admissions.filter(
            Q(application__student__first_name__icontains=name) | 
            Q(application__student__username__icontains=name)
        )

    if batch_id:
        admissions = admissions.filter(application__academic_year_id=batch_id)

    if course_id:
        admissions = admissions.filter(application__course_id=course_id)
        
    if status_filter:
        if status_filter in ['pending', 'selected', 'rejected', 'hold']:
            admissions = admissions.filter(application__status=status_filter)
        else:
            admissions = admissions.filter(status=status_filter)

    if class_id:
        admissions = admissions.filter(assigned_class_id=class_id)

    if class_year_id:
        admissions = admissions.filter(assigned_class_year_id=class_year_id)

    # Context data for filters
    batches = AcademicYear.objects.filter(institute=institute, is_active=True)
    courses = Course.objects.filter(institute=institute)
    
    # Filter classes and class years by the institute and selected academic year (batch) if provided
    classes = Class.objects.filter(institute=institute)
    if batch_id:
        classes = classes.filter(academic_year_id=batch_id)
        
    class_years = ClassYear.objects.filter(class_obj__in=classes, is_active=True)

    from django.core.paginator import Paginator
    paginator = Paginator(admissions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'institute/student_list.html', {
        'admissions': page_obj,
        'page_obj': page_obj,
        'batches': batches,
        'courses': courses,
        'classes': classes,
        'class_years': class_years,
        'admission_statuses': Admission.ADMISSION_STATUS
    })
@login_required
def update_student_status(request, admission_id):
    admission = get_object_or_404(Admission, id=admission_id, application__institute=request.user.institute)
    
    new_status = request.POST.get('status') or request.GET.get('status')
    reason = request.POST.get('reason') or request.POST.get('deletion_reason') or request.GET.get('reason', '').strip()

    if new_status in dict(Admission.ADMISSION_STATUS):
        old_status = admission.status
        admission.status = new_status
        if reason:
            admission.status_reason = reason
        admission.save()
        
        # If moving to Trash, create entry in TrashedStudent trust table for backend archiving
        if new_status == 'trashed':
            archived_info = {
                'registration_id': admission.registration_id,
                'student_name': admission.application.display_name,
                'mobile': admission.application.student.username,
                'email': admission.application.student.email,
                'course': admission.selected_course.name if admission.selected_course else None,
                'date_of_join': str(admission.date_of_join) if admission.date_of_join else None,
                'guardian_name': admission.guardian_name,
                'guardian_mobile': admission.guardian_mobile,
                'care_of': admission.care_of,
                'deletion_reason': reason,
                'deleted_by': request.user.username,
            }
            TrashedStudent.objects.create(
                admission=admission,
                registration_id=admission.registration_id or f"ADM_{admission.id}",
                student_name=admission.application.display_name or admission.application.student.username,
                mobile=admission.application.student.username,
                email=admission.application.student.email,
                course_name=admission.selected_course.name if admission.selected_course else "N/A",
                institute_name=admission.application.institute.name if admission.application.institute else "N/A",
                deletion_reason=reason or "Moved to Trash",
                deleted_by=request.user,
                archived_data=json.dumps(archived_info)
            )

        if old_status != new_status:
            from .models import log_activity
            log_activity(
                user=request.user,
                module="Admissions",
                activity=f"Updated status of student {admission.application.display_name} from {old_status} to {new_status} (Reason: {reason or 'none'})",
                institute=admission.application.institute
            )
            try:
                send_admission_status_email(admission, new_status)
            except Exception:
                pass
            
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (hasattr(request, 'is_ajax') and request.is_ajax()):
            return JsonResponse({'status': 'success', 'message': f"Student status updated to {new_status}"})

        messages.success(request, f"Status updated for {admission.application.student.first_name or admission.application.student.username}")
        return redirect(request.META.get('HTTP_REFERER', 'student_list'))
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (hasattr(request, 'is_ajax') and request.is_ajax()):
        return JsonResponse({'status': 'error', 'message': 'Invalid status'}, status=400)
    return redirect('student_list')

@login_required
def export_students_excel(request):
    institute = request.user.institute
    admissions = Admission.objects.filter(application__institute=institute).select_related('application__student', 'application__academic_year', 'application__course', 'fee_category')
    # Apply same filters as list view
    form_id = request.GET.get('form_id')
    name = request.GET.get('name')
    batch_id = request.GET.get('batch_id')
    course_id = request.GET.get('course_id')
    class_id = request.GET.get('class_id')
    class_year_id = request.GET.get('class_year_id')

    if form_id:
        admissions = admissions.filter(Q(register_number__icontains=form_id) | Q(registration_id__icontains=form_id))
    
    if name:
        admissions = admissions.filter(
            Q(application__student__first_name__icontains=name) | 
            Q(application__student__username__icontains=name)
        )
        
    if batch_id:
        admissions = admissions.filter(application__academic_year_id=batch_id)
        
    if course_id:
        admissions = admissions.filter(application__course_id=course_id)

    if class_id:
        admissions = admissions.filter(assigned_class_id=class_id)

    if class_year_id:
        admissions = admissions.filter(assigned_class_year_id=class_year_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Registry"

    headers = [
        "Admission No.", "Date of join", "Quota", "Full Name", "Academic Year", "Gender", "DOB",
        "Religion", "Caste/Community", "Category(SC/ST/OBC)", "Payment category", "Student E-mail",
        "Student Mobile No.", "Aadhar Card Number", "Course", "Division", "Class(Sem/Year)",
        # Address Details
        "Permanent House Name/No", "Permanent Location", "Permanent Post", "Permanent Pin Code", 
        "Permanent Nationality", "Permanent State", "Permanent District",
        "Communication House Name/No", "Communication Location", "Communication Post", 
        "Communication Pin Code", "Communication Nationality", "Communication State", "Communication District",
        # Family Details
        "Father Name", "Father Mobile No.", "Mother Name", "Mother Mobile", 
        "Guardian Name", "Guardian Mobile No.", "Relationship with child", "Guardian Address"
    ]
    ws.append(headers)

    def fuzzy_val(app, label_part):
        matched = ApplicationFieldValue.objects.filter(application=app, field__label__icontains=label_part).first()
        return matched.value if matched else ""

    for adm in admissions:
        app = adm.application
        row = [
            adm.register_number,
            adm.date_of_join.strftime('%Y-%m-%d') if adm.date_of_join else "",
            adm.fee_category.name if adm.fee_category else "",
            app.student.first_name,
            app.academic_year.name if app.academic_year else "",
            fuzzy_val(app, "Gender"),
            fuzzy_val(app, "DOB") or fuzzy_val(app, "Birth"),
            fuzzy_val(app, "Religion"),
            fuzzy_val(app, "Caste") or fuzzy_val(app, "Community"),
            fuzzy_val(app, "Category") or fuzzy_val(app, "Reservation"),
            adm.fee_category.name if adm.fee_category else "",
            app.student.email,
            app.student.username,
            fuzzy_val(app, "Aadhar"),
            adm.selected_course.name if adm.selected_course else app.course.name,
            fuzzy_val(app, "Division"),
            adm.joining_period.name if adm.joining_period else fuzzy_val(app, "Class"),
            # Permanent Address (Fuzzy Matching)
            fuzzy_val(app, "Permanent") and fuzzy_val(app, "House") or fuzzy_val(app, "Address"),
            fuzzy_val(app, "Permanent") and fuzzy_val(app, "Location"),
            fuzzy_val(app, "Permanent") and fuzzy_val(app, "Post"),
            fuzzy_val(app, "Permanent") and fuzzy_val(app, "Pin"),
            fuzzy_val(app, "Nationality"),
            fuzzy_val(app, "State"),
            fuzzy_val(app, "District"),
            # Communication Address
            fuzzy_val(app, "Communication") and fuzzy_val(app, "House"),
            fuzzy_val(app, "Communication") and fuzzy_val(app, "Location"),
            fuzzy_val(app, "Communication") and fuzzy_val(app, "Post"),
            fuzzy_val(app, "Communication") and fuzzy_val(app, "Pin"),
            fuzzy_val(app, "Nationality"),
            fuzzy_val(app, "State"),
            fuzzy_val(app, "District"),
            # Family
            fuzzy_val(app, "Father Name"),
            fuzzy_val(app, "Father Mobile"),
            fuzzy_val(app, "Mother Name"),
            fuzzy_val(app, "Mother Mobile"),
            adm.guardian_name,
            adm.guardian_mobile,
            adm.relationship,
            adm.guardian_address
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Students_{institute.name.replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response


# =========================
# ✅ ACADEMIC RELATIONSHIP PORTAL
# =========================

@login_required
def manage_notices(request):
    if request.user.role != 'institute' and not request.user.is_staff:
        return redirect('/')
    
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "No Institute context found.")
        return redirect('/')

    notices = NoticeBoard.objects.filter(institute=institute)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            title = request.POST.get('title')
            NoticeBoard.objects.create(
                institute=institute,
                title=title,
                content=request.POST.get('content'),
                course_id=request.POST.get('course') or None,
                assigned_class_id=request.POST.get('assigned_class') or None,
                file_attachment=request.FILES.get('file')
            )
            from .models import log_activity
            log_activity(
                user=request.user,
                module="Academic Announcements",
                activity=f"Created notice: '{title}'",
                institute=institute
            )
            messages.success(request, "Notice posted successfully.")
        elif action == 'delete':
            notice_id = request.POST.get('notice_id')
            NoticeBoard.objects.filter(id=notice_id, institute=institute).delete()
            from .models import log_activity
            log_activity(
                user=request.user,
                module="Academic Announcements",
                activity=f"Deleted notice ID: {notice_id}",
                institute=institute
            )
            messages.success(request, "Notice deleted.")
        return redirect('manage_notices')

    courses = Course.objects.filter(institute=institute)
    classes = Class.objects.filter(institute=institute)
    
    return render(request, 'institute/manage_notices.html', {
        'notices': notices,
        'courses': courses,
        'classes': classes
    })


@login_required
def manage_timetables(request):
    if request.user.role != 'institute' and not request.user.is_staff:
        return redirect('/')
        
    institute = get_current_institute(request)
    if not institute:
        return redirect('/')
    classes = Class.objects.filter(institute=institute).select_related('course', 'timetable')
    
    if request.method == 'POST':
        class_id = request.POST.get('class_id')
        image = request.FILES.get('timetable_image')
        
        if class_id and image:
            assigned_class = get_object_or_404(Class, id=class_id, institute=institute)
            timetable, created = Timetable.objects.get_or_create(assigned_class=assigned_class)
            timetable.image_file = image
            timetable.save()
            from .models import log_activity
            log_activity(
                user=request.user,
                module="Academic Announcements",
                activity=f"Uploaded timetable for class: {assigned_class.name}",
                institute=institute
            )
            messages.success(request, f"Timetable updated for {assigned_class.name}")
        return redirect('manage_timetables')
        
    return render(request, 'institute/manage_timetables.html', {'classes': classes})


@login_required
def enter_academic_results(request):
    if request.user.role != 'institute' and not request.user.is_staff:
        return redirect('/')
        
    institute = get_current_institute(request)
    if not institute:
        return redirect('/')
    
    classes = Class.objects.filter(institute=institute)
    periods = CourseSubCategory.objects.all()
    
    selected_class = None
    selected_period = None
    selected_subject = None
    students = []
    subjects = []

    if request.GET.get('class_id'):
        selected_class = Class.objects.filter(id=request.GET.get('class_id')).first()
        if selected_class:
            subjects = selected_class.subjects.all()
            if not subjects.exists():
                subjects = Subject.objects.filter(
                    Q(classes=selected_class) | Q(course=selected_class.course)
                ).distinct()
            students = Admission.objects.filter(
                Q(assigned_class=selected_class) | Q(assigned_class__isnull=True, selected_course=selected_class.course)
            ).exclude(status='trashed').select_related('application__student')

    if request.GET.get('period_id'):
        selected_period = get_object_or_404(CourseSubCategory, id=request.GET.get('period_id'))

    if request.GET.get('subject_id'):
        selected_subject = get_object_or_404(Subject, id=request.GET.get('subject_id'), institute=institute)

    if request.method == 'POST':
        # Bulk save results
        subject_id = request.POST.get('subject_id')
        period_id = request.POST.get('period_id')
        
        for key, value in request.POST.items():
            if key.startswith('marks_'):
                admission_id = key.replace('marks_', '')
                marks = value
                remarks = request.POST.get(f'remarks_{admission_id}')
                
                if marks:
                    admission = Admission.objects.get(id=admission_id)
                    AcademicResult.objects.update_or_create(
                        admission=admission,
                        subject_id=subject_id,
                        period_id=period_id,
                        defaults={
                            'marks_obtained': marks,
                            'remarks': remarks
                        }
                    )
        messages.success(request, "Academic results saved successfully.")
        from .models import log_activity
        log_activity(
            user=request.user,
            module="Academics",
            activity=f"Saved academic results for class: {selected_class.name if selected_class else 'n/a'}, subject ID: {subject_id}, period ID: {period_id}",
            institute=institute
        )
        return redirect(f"{request.path}?class_id={selected_class.id}&period_id={period_id}&subject_id={subject_id}")

    return render(request, 'institute/enter_results.html', {
        'classes': classes,
        'periods': periods,
        'subjects': subjects,
        'students': students,
        'selected_class': selected_class,
        'selected_period': selected_period,
        'selected_subject': selected_subject
    })

# =========================
# PAYMENT DETAILS
# =========================
from django.core.paginator import Paginator
from applications.models import Payment

@login_required
def payment_list_view(request):
    institute = getattr(request.user, 'institute', None)
    if not institute:
        return HttpResponse("Unauthorized", status=401)
        
    # Filter payments for this institute
    payments = Payment.objects.filter(application__institute=institute).select_related(
        'application__student', 'application__course'
    ).order_by('-created_at')
    
    # Search and Filter
    search_query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', 'all')
    
    if search_query:
        payments = payments.filter(
            Q(application__student__first_name__icontains=search_query) |
            Q(application__student__username__icontains=search_query) |
            Q(gateway_transaction_id__icontains=search_query) |
            Q(id__icontains=search_query)
        )
        
    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)
    if status_filter and status_filter != 'all':
        payments = payments.filter(status=status_filter)
        
    # Pagination
    paginator = Paginator(payments, 10) # 10 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'institute/payment_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter
    })

@login_required
def export_payments_excel(request):
    institute = getattr(request.user, 'institute', None)
    if not institute:
        return HttpResponse("Unauthorized", status=401)
        
    payments = Payment.objects.filter(application__institute=institute).select_related('application__student')
    
    search_query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', 'all')
    
    if search_query:
        payments = payments.filter(
            Q(application__student__first_name__icontains=search_query) |
            Q(application__student__username__icontains=search_query) |
            Q(gateway_transaction_id__icontains=search_query) |
            Q(application__id__icontains=search_query)
        ).distinct()

    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)
    if status_filter and status_filter != 'all':
        payments = payments.filter(status=status_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Payments"

    headers = ['Form Number', 'Student Name', 'Mobile No.', 'Payment Status', 'Amount', 'Payment Mode', 'Gateway Transaction ID', 'Payment Date']
    ws.append(headers)

    for payment in payments:
        student_name = payment.application.display_name
        student_mobile = payment.application.student_mobile
        created_at_str = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else (payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else '')
        status_display = dict(Payment._meta.get_field('status').choices).get(payment.status, payment.status).title()
        
        ws.append([
            str(payment.application.id),
            student_name,
            student_mobile,
            status_display,
            float(payment.amount),
            payment.payment_mode or '-',
            payment.gateway_transaction_id or '-',
            created_at_str
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Payment_Details_{institute.name}.xlsx"'
    wb.save(response)
    return response


# Helper to convert numbers to words (Rupees and Paise)
def num_to_words(n):
    try:
        n = float(n)
    except (ValueError, TypeError):
        return ""
    
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def _convert_below_thousand(num):
        word = ""
        if num >= 100:
            word += units[num // 100] + " Hundred "
            num %= 100
        if num >= 20:
            word += tens[num // 10] + " "
            num %= 10
        if num > 0:
            word += units[num] + " "
        return word

    rupees = int(n)
    paise = int(round((n - rupees) * 100))

    if rupees == 0:
        words = "Zero"
    else:
        words = ""
        cr = rupees // 10000000
        rupees %= 10000000
        lakh = rupees // 100000
        rupees %= 100000
        thousand = rupees // 1000
        rupees %= 1000
        hundreds = rupees

        if cr > 0:
            words += _convert_below_thousand(cr).strip() + " Crore "
        if lakh > 0:
            words += _convert_below_thousand(lakh).strip() + " Lakh "
        if thousand > 0:
            words += _convert_below_thousand(thousand).strip() + " Thousand "
        if hundreds > 0:
            words += _convert_below_thousand(hundreds).strip() + " "

    res = words.strip() + " Rupees"
    if paise > 0:
        res += " and " + _convert_below_thousand(paise).strip() + " Paise"
    res += " Only"
    return res


def generate_receipt_number():
    import re
    payments_with_rcpt = StudentFeePayment.objects.filter(receipt_number__isnull=False).exclude(receipt_number='')
    max_num = 32210
    for p in payments_with_rcpt:
        nums = re.findall(r'\d+', p.receipt_number or '')
        if nums:
            val = int(nums[-1])
            if val >= max_num:
                max_num = val + 1
    if max_num == 32210:
        count = StudentFeePayment.objects.count()
        max_num = 32210 + count
    return str(max_num)


def ensure_receipt_numbers():
    unassigned = StudentFeePayment.objects.filter(Q(receipt_number__isnull=True) | Q(receipt_number=''))
    if unassigned.exists():
        start_no = 32200
        for p in unassigned:
            if not p.receipt_number:
                same_batch = StudentFeePayment.objects.filter(
                    admission=p.admission,
                    payment_date=p.payment_date,
                    reference_no=p.reference_no,
                    created_at__gte=p.created_at - datetime.timedelta(seconds=5),
                    created_at__lte=p.created_at + datetime.timedelta(seconds=5)
                ).filter(receipt_number__isnull=False).first()
                if same_batch and same_batch.receipt_number:
                    p.receipt_number = same_batch.receipt_number
                else:
                    p.receipt_number = str(start_no + p.id)
                p.save()


def group_payments_by_receipt(payments_qs):
    ensure_receipt_numbers()
    receipt_dict = {}
    receipt_order = []
    
    for payment in payments_qs:
        rcpt_id = payment.receipt_number or f"REC-{payment.id}"
        if rcpt_id not in receipt_dict:
            receipt_dict[rcpt_id] = {
                'receipt_number': rcpt_id,
                'admission': payment.admission,
                'payment_date': payment.payment_date,
                'payment_mode': payment.payment_mode,
                'payment_mode_display': payment.get_payment_mode_display(),
                'reference_no': payment.reference_no,
                'remarks': payment.remarks,
                'is_cancelled': payment.is_cancelled,
                'cancelled_at': payment.cancelled_at,
                'cancellation_reason': payment.cancellation_reason,
                'created_at': payment.created_at,
                'payments': [],
                'fee_items': [],
                'total_amount_paid': 0.0,
                'total_fine_paid': 0.0,
                'grand_total': 0.0,
            }
            receipt_order.append(rcpt_id)
        
        g = receipt_dict[rcpt_id]
        if payment.is_cancelled:
            g['is_cancelled'] = True
            g['cancelled_at'] = payment.cancelled_at
            g['cancellation_reason'] = payment.cancellation_reason

        g['payments'].append(payment)
        fee_name = payment.fee_head.fee_type.name
        g['fee_items'].append({
            'name': fee_name,
            'amount_paid': float(payment.amount_paid),
            'fine_paid': float(payment.fine_paid),
            'total': float(payment.amount_paid) + float(payment.fine_paid)
        })
        g['total_amount_paid'] += float(payment.amount_paid)
        g['total_fine_paid'] += float(payment.fine_paid)
        g['grand_total'] += (float(payment.amount_paid) + float(payment.fine_paid))
        if payment.reference_no and not g['reference_no']:
            g['reference_no'] = payment.reference_no
        if payment.remarks and not g['remarks']:
            g['remarks'] = payment.remarks

    grouped_list = []
    for rcpt_id in receipt_order:
        g = receipt_dict[rcpt_id]
        g['fee_types_str'] = ", ".join(item['name'] for item in g['fee_items'])
        grouped_list.append(g)
        
    return grouped_list


@login_required
def receipt_list(request):
    institute = request.user.institute
    from django.utils import timezone
    
    academic_years = AcademicYear.objects.filter(institute=institute, is_active=True).order_by('-name')
    courses = Course.objects.filter(institute=institute).order_by('name')
    categories = CourseCategory.objects.all().order_by('name')
    
    selected_year = request.GET.get('year', '')
    selected_course = request.GET.get('course', '')
    selected_category = request.GET.get('category', '')
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('q', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    payments = StudentFeePayment.objects.filter(
        admission__application__institute=institute
    ).select_related(
        'admission', 'admission__application', 'admission__application__student',
        'admission__selected_course', 'admission__assigned_class', 'fee_head', 'fee_head__fee_type'
    ).order_by('-created_at', '-id')
    
    if selected_year:
        payments = payments.filter(admission__application__academic_year_id=selected_year)
    if selected_course:
        payments = payments.filter(admission__selected_course_id=selected_course)
    if selected_category:
        payments = payments.filter(admission__selected_course__category_id=selected_category)
    if status_filter == 'active':
        payments = payments.filter(is_cancelled=False)
    elif status_filter == 'cancelled':
        payments = payments.filter(is_cancelled=True)
        
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
        
    if search_query:
        payments = payments.filter(
            Q(receipt_number__icontains=search_query) |
            Q(admission__registration_id__icontains=search_query) |
            Q(admission__application__form_no__icontains=search_query) |
            Q(admission__application__student__first_name__icontains=search_query) |
            Q(admission__application__student__last_name__icontains=search_query) |
            Q(reference_no__icontains=search_query)
        )
        
    grouped = group_payments_by_receipt(payments)
    
    paginator = Paginator(grouped, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'receipts': page_obj.object_list,
        'page_obj': page_obj,
        'academic_years': academic_years,
        'courses': courses,
        'categories': categories,
        'selected_year': selected_year,
        'selected_course': selected_course,
        'selected_category': selected_category,
        'status_filter': status_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'institute/receipt_list.html', context)


@login_required
def cancel_fee_receipt(request, receipt_number):
    institute = request.user.institute
    from django.utils import timezone
    if request.method == 'POST':
        reason = request.POST.get('cancellation_reason', 'Cancelled by administrator').strip()
        payments = StudentFeePayment.objects.filter(
            receipt_number=receipt_number,
            admission__application__institute=institute
        )
        if payments.exists():
            payments.update(
                is_cancelled=True,
                cancelled_at=timezone.now(),
                cancellation_reason=reason
            )
            messages.success(request, f"Receipt #{receipt_number} has been cancelled successfully.")
        else:
            messages.error(request, f"Receipt #{receipt_number} not found.")
            
    next_url = request.POST.get('next', request.META.get('HTTP_REFERER', '/institute/receipts/'))
    return redirect(next_url)


# =============================================================================
# STUDENT FEE MANAGEMENT VIEWS
# =============================================================================
@login_required
def manage_student_fees(request, admission_id):
    institute = request.user.institute
    admission = get_object_or_404(Admission, id=admission_id, application__institute=institute)
    
    if request.method == 'POST':
        class_year_id = request.POST.get('class_year_id')
        fee_category_id = request.POST.get('fee_category_id')
        custom_discount = request.POST.get('custom_discount')
        if class_year_id:
            admission.assigned_class_year_id = class_year_id
        else:
            admission.assigned_class_year = None
            
        if fee_category_id:
            admission.assigned_fee_category_id = fee_category_id
        else:
            admission.assigned_fee_category = None
            
        if custom_discount:
            try:
                admission.custom_discount_percentage = float(custom_discount)
            except ValueError:
                pass
        else:
            admission.custom_discount_percentage = None
            
        admission.save()
        messages.success(request, f"Fee assignment updated for {admission.application.display_name}")
        return redirect('manage_student_fees', admission_id=admission_id)
        
    # Dropdown fallback logic if class is optional/blank (pull course-level class years)
    if admission.assigned_class:
        class_years = ClassYear.objects.filter(class_obj=admission.assigned_class, is_active=True)
    else:
        class_years = ClassYear.objects.filter(class_obj__course=admission.selected_course, is_active=True)
    fee_categories = FeeCategoryMaster.objects.filter(is_active=True)
    
    # Mapped Fee Structure
    fee_heads_data = []
    total_demand = 0.0
    total_collected = 0.0
    total_pending = 0.0
    total_fine = 0.0
    
    # Active Class Resolution fallback
    active_class = admission.assigned_class or (admission.assigned_class_year.class_obj if admission.assigned_class_year else None)
    
    if active_class and admission.assigned_class_year and admission.assigned_fee_category:
        structure = FeeStructure.objects.filter(
            academic_year=admission.application.academic_year,
            institute=institute,
            course=admission.selected_course,
            class_obj=active_class,
            class_year=admission.assigned_class_year,
            fee_category=admission.assigned_fee_category
        ).first()
        
        if structure:
            heads = structure.heads.filter(is_active=True)
            for head in heads:
                # CHECK IF DISCOUNT IS ELIGIBLE (User Enhancement 1)
                if head.fee_type.is_discountable:
                    discount_pct = admission.custom_discount_percentage if admission.custom_discount_percentage is not None else admission.assigned_fee_category.discount_percentage
                    discount_pct = discount_pct or 0
                else:
                    discount_pct = 0
                
                discounted_amount = float(head.amount) * (1 - float(discount_pct) / 100)
                
                has_fine = False
                fine_amt = 0.0
                if datetime.date.today() > head.due_date:
                    has_fine = True
                    fine_amt = float(head.fine_amount)
                    
                payments = StudentFeePayment.objects.filter(admission=admission, fee_head=head, is_cancelled=False)
                paid_fee = sum(float(p.amount_paid) for p in payments)
                paid_fine = sum(float(p.fine_paid) for p in payments)
                
                pending_fee = discounted_amount - paid_fee
                pending_fine = fine_amt - paid_fine if has_fine else 0.0
                
                total_for_head = discounted_amount + fine_amt
                total_paid_for_head = paid_fee + paid_fine
                total_pending_for_head = pending_fee + pending_fine
                
                fee_heads_data.append({
                    'head': head,
                    'discounted_amount': discounted_amount,
                    'discount_pct': discount_pct,
                    'fine_amount': fine_amt,
                    'paid_fee': paid_fee,
                    'paid_fine': paid_fine,
                    'pending_fee': max(0.0, pending_fee),
                    'pending_fine': max(0.0, pending_fine),
                    'total': total_for_head,
                    'total_paid': total_paid_for_head,
                    'total_pending': max(0.0, total_pending_for_head)
                })
                
                total_demand += discounted_amount
                total_collected += paid_fee
                total_pending += max(0.0, pending_fee)
                total_fine += paid_fine
                
    raw_history = StudentFeePayment.objects.filter(admission=admission).select_related('fee_head__fee_type').order_by('-created_at')
    payments_history = group_payments_by_receipt(raw_history)
    
    return render(request, 'institute/manage_student_fees.html', {
        'admission': admission,
        'class_years': class_years,
        'fee_categories': fee_categories,
        'fee_heads_data': fee_heads_data,
        'total_demand': total_demand,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'total_fine_collected': total_fine,
        'payments_history': payments_history
    })


@login_required
def collect_student_fee(request, admission_id, head_id):
    admission = get_object_or_404(Admission, id=admission_id, application__institute=request.user.institute)
    head = get_object_or_404(FeeHead, id=head_id)
    
    if request.method == 'POST':
        amount_paid = request.POST.get('amount_paid', 0)
        fine_paid = request.POST.get('fine_paid', 0)
        payment_mode = request.POST.get('payment_mode', 'cash')
        reference_no = request.POST.get('reference_no', '')
        remarks = request.POST.get('remarks', '')
        
        try:
            amt_val = float(amount_paid) if amount_paid else 0.0
            fine_val = float(fine_paid) if fine_paid else 0.0
            rcpt_no = generate_receipt_number()
            
            StudentFeePayment.objects.create(
                admission=admission,
                fee_head=head,
                amount_paid=amt_val,
                fine_paid=fine_val,
                payment_mode=payment_mode,
                reference_no=reference_no,
                receipt_number=rcpt_no,
                remarks=remarks,
                payment_date=datetime.date.today()
            )
            
            from .models import log_activity
            log_activity(
                user=request.user,
                module="Fee Management",
                activity=f"Collected ₹{amt_val:.2f} (Fine: ₹{fine_val:.2f}) for student {admission.application.display_name} (Fee Head: {head.fee_type.name})",
                institute=admission.application.institute
            )
            
            messages.success(request, f"Payment of ₹{amt_val + fine_val} recorded successfully!")
        except Exception as e:
            messages.error(request, f"Error processing payment: {str(e)}")
            
    return redirect('manage_student_fees', admission_id=admission_id)


@login_required
def print_fee_receipt(request, receipt_number):
    institute = getattr(request.user, 'institute', None)
    if not institute and request.user.is_staff:
        institute = Institute.objects.first()

    payments = StudentFeePayment.objects.filter(
        receipt_number=receipt_number,
        admission__application__institute=institute
    ).select_related('admission__application__student', 'admission__selected_course', 'admission__assigned_class', 'admission__assigned_class_year', 'fee_head__fee_type')

    if not payments.exists():
        messages.error(request, f"Receipt #{receipt_number} not found.")
        return redirect('student_list')

    first_payment = payments.first()
    admission = first_payment.admission
    app = admission.application

    fee_items = []
    grand_total = 0.0

    for p in payments:
        amt = float(p.amount_paid)
        if amt > 0:
            fee_items.append({
                'particulars': p.fee_head.fee_type.name,
                'amount': amt,
                'rs': int(amt),
                'ps': int(round((amt - int(amt)) * 100))
            })
            grand_total += amt

        fine = float(p.fine_paid)
        if fine > 0:
            fee_items.append({
                'particulars': f"Fine ({p.fee_head.fee_type.name})",
                'amount': fine,
                'rs': int(fine),
                'ps': int(round((fine - int(fine)) * 100))
            })
            grand_total += fine

    total_rs = int(grand_total)
    total_ps = int(round((grand_total - total_rs) * 100))
    amount_in_words = num_to_words(grand_total)

    class_name = admission.assigned_class_year.name if admission.assigned_class_year else (admission.assigned_class.name if admission.assigned_class else "Semester / Class")
    course_name = admission.selected_course.name if admission.selected_course else (app.course.name if app.course else "")

    context = {
        'receipt_number': receipt_number,
        'payments': payments,
        'first_payment': first_payment,
        'admission': admission,
        'app': app,
        'institute': institute,
        'fee_items': fee_items,
        'grand_total': grand_total,
        'total_rs': total_rs,
        'total_ps': f"{total_ps:02d}",
        'amount_in_words': amount_in_words,
        'class_name': class_name,
        'course_name': course_name,
        'payment_date': first_payment.payment_date,
        'payment_mode': first_payment.get_payment_mode_display(),
        'reference_no': first_payment.reference_no,
    }
    return render(request, 'institute/print_receipt.html', context)


def clean_id_param(val):
    if not val or val in ['None', 'null', 'undefined', '']:
        return None
    return val


@login_required
def fee_reports(request):
    institute = request.user.institute
    
    academic_year_id = clean_id_param(request.GET.get('academic_year_id'))
    course_id = clean_id_param(request.GET.get('course_id'))
    class_id = clean_id_param(request.GET.get('class_id'))
    class_year_id = clean_id_param(request.GET.get('class_year_id'))
    fee_category_id = clean_id_param(request.GET.get('fee_category_id'))
    fee_type_id = clean_id_param(request.GET.get('fee_type_id'))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    admissions_qs = Admission.objects.filter(application__institute=institute)
    if academic_year_id:
        admissions_qs = admissions_qs.filter(application__academic_year_id=academic_year_id)
    if course_id:
        admissions_qs = admissions_qs.filter(selected_course_id=course_id)
    if class_id:
        admissions_qs = admissions_qs.filter(assigned_class_id=class_id)
    if class_year_id:
        admissions_qs = admissions_qs.filter(assigned_class_year_id=class_year_id)
    if fee_category_id:
        admissions_qs = admissions_qs.filter(assigned_fee_category_id=fee_category_id)
    if date_from:
        admissions_qs = admissions_qs.filter(date_of_join__gte=date_from)
    if date_to:
        admissions_qs = admissions_qs.filter(date_of_join__lte=date_to)
        
    admissions = list(admissions_qs.select_related('application__student', 'assigned_class', 'assigned_class_year', 'assigned_fee_category'))
    
    roster_data = []
    total_all_demand = 0.0
    total_all_collected = 0.0
    total_all_pending = 0.0
    
    for adm in admissions:
        active_class = adm.assigned_class or (adm.assigned_class_year.class_obj if adm.assigned_class_year else None)
        if active_class and adm.assigned_class_year and adm.assigned_fee_category:
            structure = FeeStructure.objects.filter(
                academic_year=adm.application.academic_year,
                institute=institute,
                course=adm.selected_course,
                class_obj=active_class,
                class_year=adm.assigned_class_year,
                fee_category=adm.assigned_fee_category
            ).first()
            
            if structure:
                heads = structure.heads.filter(is_active=True)
                if fee_type_id:
                    heads = heads.filter(fee_type_id=fee_type_id)
                    
                std_demand = 0.0
                std_collected = 0.0
                std_fines = 0.0
                
                for head in heads:
                    if head.fee_type.is_discountable:
                        discount_pct = adm.custom_discount_percentage if adm.custom_discount_percentage is not None else adm.assigned_fee_category.discount_percentage
                        discount_pct = discount_pct or 0
                    else:
                        discount_pct = 0
                        
                    discounted_amount = float(head.amount) * (1 - float(discount_pct) / 100)
                    std_demand += discounted_amount
                    
                    payments = StudentFeePayment.objects.filter(admission=adm, fee_head=head, is_cancelled=False)
                    std_collected += sum(float(p.amount_paid) for p in payments)
                    std_fines += sum(float(p.fine_paid) for p in payments)
                    
                pending_fee = std_demand - std_collected
                
                roster_data.append({
                    'admission': adm,
                    'fee_category': adm.assigned_fee_category,
                    'discount_pct': discount_pct,
                    'demand': std_demand,
                    'collected': std_collected,
                    'fine_collected': std_fines,
                    'pending': max(0.0, pending_fee),
                })
                
                total_all_demand += std_demand
                total_all_collected += std_collected + std_fines
                total_all_pending += max(0.0, pending_fee)
                
    ledger_payments = StudentFeePayment.objects.filter(admission__application__institute=institute).select_related('admission__application__student', 'fee_head__fee_type').order_by('-created_at')
    
    if fee_type_id:
        ledger_payments = ledger_payments.filter(fee_head__fee_type_id=fee_type_id)
    if fee_category_id:
        ledger_payments = ledger_payments.filter(admission__assigned_fee_category_id=fee_category_id)
    if date_from:
        ledger_payments = ledger_payments.filter(payment_date__gte=date_from)
    if date_to:
        ledger_payments = ledger_payments.filter(payment_date__lte=date_to)
        
    grouped_ledger = group_payments_by_receipt(ledger_payments)

    batches = AcademicYear.objects.filter(institute=institute, is_active=True)
    courses = Course.objects.filter(institute=institute)
    classes = Class.objects.filter(institute=institute)
    class_years = ClassYear.objects.filter(is_active=True)
    fee_categories = FeeCategoryMaster.objects.filter(is_active=True)
    fee_types = FeeType.objects.filter(is_active=True)
    
    return render(request, 'institute/fee_reports.html', {
        'roster_data': roster_data,
        'total_all_demand': total_all_demand,
        'total_all_collected': total_all_collected,
        'total_all_pending': total_all_pending,
        'ledger_payments': grouped_ledger[:100],
        'batches': batches,
        'courses': courses,
        'classes': classes,
        'class_years': class_years,
        'fee_categories': fee_categories,
        'fee_types': fee_types,
        'selected_year': academic_year_id,
        'selected_course': course_id,
        'selected_class': class_id,
        'selected_class_year': class_year_id,
        'selected_fee_category': fee_category_id,
        'selected_fee_type': fee_type_id,
        'date_from': date_from,
        'date_to': date_to
    })


@login_required
def collect_multiple_fees(request, admission_id):
    admission = get_object_or_404(Admission, id=admission_id, application__institute=request.user.institute)
    
    if request.method == 'POST':
        head_ids = request.POST.getlist('head_ids')
        payment_mode = request.POST.get('payment_mode', 'cash')
        reference_no = request.POST.get('reference_no', '')
        remarks = request.POST.get('remarks', '')
        
        if not head_ids:
            messages.error(request, "No fee heads were selected for collection.")
            return redirect('manage_student_fees', admission_id=admission_id)
            
        success_count = 0
        total_collected = 0.0
        
        try:
            active_class = admission.assigned_class or (admission.assigned_class_year.class_obj if admission.assigned_class_year else None)
            
            if active_class and admission.assigned_class_year and admission.assigned_fee_category:
                structure = FeeStructure.objects.filter(
                    academic_year=admission.application.academic_year,
                    institute=request.user.institute,
                    course=admission.selected_course,
                    class_obj=active_class,
                    class_year=admission.assigned_class_year,
                    fee_category=admission.assigned_fee_category
                ).first()
                
                if structure:
                    heads = structure.heads.filter(id__in=head_ids, is_active=True)
                    rcpt_no = generate_receipt_number() # Single receipt ID for all items collected together!
                    
                    for head in heads:
                        if head.fee_type.is_discountable:
                            discount_pct = admission.custom_discount_percentage if admission.custom_discount_percentage is not None else admission.assigned_fee_category.discount_percentage
                            discount_pct = discount_pct or 0
                        else:
                            discount_pct = 0
                            
                        discounted_amount = float(head.amount) * (1 - float(discount_pct) / 100)
                        
                        has_fine = False
                        fine_amt = 0.0
                        if datetime.date.today() > head.due_date:
                            has_fine = True
                            fine_amt = float(head.fine_amount)
                            
                        payments = StudentFeePayment.objects.filter(admission=admission, fee_head=head, is_cancelled=False)
                        paid_fee = sum(float(p.amount_paid) for p in payments)
                        paid_fine = sum(float(p.fine_paid) for p in payments)
                        
                        pending_fee = discounted_amount - paid_fee
                        pending_fine = fine_amt - paid_fine if has_fine else 0.0
                        
                        if pending_fee > 0 or pending_fine > 0:
                            amt_val = max(0.0, pending_fee)
                            fine_val = max(0.0, pending_fine)
                            
                            StudentFeePayment.objects.create(
                                admission=admission,
                                fee_head=head,
                                amount_paid=amt_val,
                                fine_paid=fine_val,
                                payment_mode=payment_mode,
                                reference_no=reference_no,
                                receipt_number=rcpt_no,
                                remarks=remarks,
                                payment_date=datetime.date.today()
                            )
                            total_collected += (amt_val + fine_val)
                            success_count += 1
                            
            if success_count > 0:
                from .models import log_activity
                log_activity(
                    user=request.user,
                    module="Fee Management",
                    activity=f"Bulk collected ₹{total_collected:.2f} for {success_count} fee heads of student {admission.application.display_name} (Receipt #{rcpt_no})",
                    institute=admission.application.institute
                )
                messages.success(request, f"Successfully collected ₹{total_collected:.2f} for {success_count} fee heads! (Receipt #{rcpt_no})")
            else:
                messages.warning(request, "Selected fee heads have already been fully paid.")
        except Exception as e:
            messages.error(request, f"Error processing bulk collection: {str(e)}")
            
    return redirect('manage_student_fees', admission_id=admission_id)


# =========================
# SYSTEM BACKUP & RESTORE
# =========================
from io import BytesIO
from django.contrib.admin.views.decorators import staff_member_required
from .backup import generate_zip_backup

@staff_member_required
def system_backup_view(request):
    """
    Renders the backup dashboard template where staff can request a backup.
    """
    return render(request, 'institute/backup.html')

@staff_member_required
def download_backup_view(request):
    """
    Generates and streams a zip file backup.
    """
    try:
        buffer = BytesIO()
        generate_zip_backup(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        filename = f"jdt_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        from .models import log_activity
        log_activity(
            user=request.user,
            module="System Backup",
            activity=f"Downloaded system ZIP backup: {filename}",
            institute=None
        )
        return response
    except Exception as e:
        messages.error(request, f"Failed to generate backup: {str(e)}")
        return redirect('system_backup')


# =============================================================================
# EXPORT FEE REPORTS AND AUDITS
# =============================================================================
@login_required
def export_fee_reports_excel(request):
    institute = request.user.institute
    report_type = request.GET.get('report_type', 'entire')
    
    academic_year_id = clean_id_param(request.GET.get('academic_year_id'))
    course_id = clean_id_param(request.GET.get('course_id'))
    class_id = clean_id_param(request.GET.get('class_id'))
    class_year_id = clean_id_param(request.GET.get('class_year_id'))
    fee_category_id = clean_id_param(request.GET.get('fee_category_id'))
    fee_type_id = clean_id_param(request.GET.get('fee_type_id'))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Filter admissions (Roster)
    admissions_qs = Admission.objects.filter(application__institute=institute)
    if academic_year_id:
        admissions_qs = admissions_qs.filter(application__academic_year_id=academic_year_id)
    if course_id:
        admissions_qs = admissions_qs.filter(selected_course_id=course_id)
    if class_id:
        admissions_qs = admissions_qs.filter(assigned_class_id=class_id)
    if class_year_id:
        admissions_qs = admissions_qs.filter(assigned_class_year_id=class_year_id)
    if fee_category_id:
        admissions_qs = admissions_qs.filter(assigned_fee_category_id=fee_category_id)
    if date_from:
        admissions_qs = admissions_qs.filter(date_of_join__gte=date_from)
    if date_to:
        admissions_qs = admissions_qs.filter(date_of_join__lte=date_to)
        
    admissions = admissions_qs.select_related('application__student', 'assigned_class', 'assigned_class_year', 'assigned_fee_category')
    
    # Filter payments (Audit Log)
    ledger_payments = StudentFeePayment.objects.filter(admission__application__institute=institute).select_related('admission__application__student', 'fee_head__fee_type').order_by('-created_at')
    if fee_type_id:
        ledger_payments = ledger_payments.filter(fee_head__fee_type_id=fee_type_id)
    if fee_category_id:
        ledger_payments = ledger_payments.filter(admission__assigned_fee_category_id=fee_category_id)
    if date_from:
        ledger_payments = ledger_payments.filter(payment_date__gte=date_from)
    if date_to:
        ledger_payments = ledger_payments.filter(payment_date__lte=date_to)

    wb = Workbook()
    ws = wb.active
    
    if report_type == 'audit':
        ws.title = "Collection Audit Log"
        headers = ['Receipt ID', 'Date', 'Candidate Name', 'Register No', 'Demanded Heads', 'Amount Paid', 'Penalty Paid', 'Mode', 'Reference No']
        ws.append(headers)
        
        grouped_audit = group_payments_by_receipt(ledger_payments)
        for p in grouped_audit:
            ws.append([
                f"#{p['receipt_number']}",
                p['payment_date'].strftime('%Y-%m-%d') if p['payment_date'] else '-',
                p['admission'].application.display_name,
                p['admission'].registration_id or '-',
                p['fee_types_str'],
                float(p['total_amount_paid']),
                float(p['total_fine_paid']),
                p['payment_mode_display'],
                p['reference_no'] or '-'
            ])
            
    else:
        # Entire Fee Report or Outstanding Balance Roster
        if report_type == 'roster':
            ws.title = "Outstanding Balance Roster"
        else:
            ws.title = "Entire Fee Report"
            
        headers = ['Register No', 'Student Name', 'Term & Class', 'Fee Category', 'Net Demand', 'Paid Fees', 'Fines Collected', 'Pending Balance']
        ws.append(headers)
        
        for adm in admissions:
            active_class = adm.assigned_class or (adm.assigned_class_year.class_obj if adm.assigned_class_year else None)
            if active_class and adm.assigned_class_year and adm.assigned_fee_category:
                structure = FeeStructure.objects.filter(
                    academic_year=adm.application.academic_year,
                    institute=institute,
                    course=adm.selected_course,
                    class_obj=active_class,
                    class_year=adm.assigned_class_year,
                    fee_category=adm.assigned_fee_category
                ).first()
                
                if structure:
                    heads = structure.heads.filter(is_active=True)
                    if fee_type_id:
                        heads = heads.filter(fee_type_id=fee_type_id)
                        
                    std_demand = 0.0
                    std_collected = 0.0
                    std_fines = 0.0
                    
                    for head in heads:
                        if head.fee_type.is_discountable:
                            discount_pct = adm.custom_discount_percentage if adm.custom_discount_percentage is not None else adm.assigned_fee_category.discount_percentage
                            discount_pct = discount_pct or 0
                        else:
                            discount_pct = 0
                            
                        discounted_amount = float(head.amount) * (1 - float(discount_pct) / 100)
                        std_demand += discounted_amount
                        
                        payments = StudentFeePayment.objects.filter(admission=adm, fee_head=head, is_cancelled=False)
                        std_collected += sum(float(p.amount_paid) for p in payments)
                        std_fines += sum(float(p.fine_paid) for p in payments)
                        
                    pending_fee = std_demand - std_collected
                    pending_fee = max(0.0, pending_fee)
                    
                    if report_type == 'roster' and pending_fee <= 0:
                        continue # Skip fully paid students for outstanding roster
                        
                    class_term_str = f"{adm.assigned_class_year.name} - {adm.assigned_class.name}"
                    ws.append([
                        adm.registration_id or '-',
                        adm.application.display_name,
                        class_term_str,
                        adm.assigned_fee_category.name,
                        std_demand,
                        std_collected,
                        std_fines,
                        pending_fee
                    ])
                    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Fee_Report_{report_type}_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_fee_reports_pdf(request):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    
    institute = request.user.institute
    report_type = request.GET.get('report_type', 'entire')
    
    academic_year_id = clean_id_param(request.GET.get('academic_year_id'))
    course_id = clean_id_param(request.GET.get('course_id'))
    class_id = clean_id_param(request.GET.get('class_id'))
    class_year_id = clean_id_param(request.GET.get('class_year_id'))
    fee_category_id = clean_id_param(request.GET.get('fee_category_id'))
    fee_type_id = clean_id_param(request.GET.get('fee_type_id'))
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Filter admissions (Roster)
    admissions_qs = Admission.objects.filter(application__institute=institute)
    if academic_year_id:
        admissions_qs = admissions_qs.filter(application__academic_year_id=academic_year_id)
    if course_id:
        admissions_qs = admissions_qs.filter(selected_course_id=course_id)
    if class_id:
        admissions_qs = admissions_qs.filter(assigned_class_id=class_id)
    if class_year_id:
        admissions_qs = admissions_qs.filter(assigned_class_year_id=class_year_id)
    if fee_category_id:
        admissions_qs = admissions_qs.filter(assigned_fee_category_id=fee_category_id)
    if date_from:
        admissions_qs = admissions_qs.filter(date_of_join__gte=date_from)
    if date_to:
        admissions_qs = admissions_qs.filter(date_of_join__lte=date_to)
        
    admissions = admissions_qs.select_related('application__student', 'assigned_class', 'assigned_class_year', 'assigned_fee_category')
    
    # Filter payments (Audit Log)
    ledger_payments = StudentFeePayment.objects.filter(admission__application__institute=institute).select_related('admission__application__student', 'fee_head__fee_type').order_by('-created_at')
    if fee_type_id:
        ledger_payments = ledger_payments.filter(fee_head__fee_type_id=fee_type_id)
    if fee_category_id:
        ledger_payments = ledger_payments.filter(admission__assigned_fee_category_id=fee_category_id)
    if date_from:
        ledger_payments = ledger_payments.filter(payment_date__gte=date_from)
    if date_to:
        ledger_payments = ledger_payments.filter(payment_date__lte=date_to)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold', leading=20, alignment=1, spaceAfter=15)
    header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.whitesmoke)
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, fontName='Helvetica', leading=10)
    
    # Add Title
    title_text = "ENTIRE FEE REPORT"
    if report_type == 'roster':
        title_text = "OUTSTANDING BALANCE ROSTER"
    elif report_type == 'audit':
        title_text = "DAILY COLLECTION AUDIT LOG"
        
    elements.append(Paragraph(f"{institute.name.upper()}<br/><font size='12'>{title_text}</font>", title_style))
    elements.append(Spacer(1, 10))
    
    if report_type == 'audit':
        headers = [
            Paragraph("Receipt ID", header_style),
            Paragraph("Date", header_style),
            Paragraph("Candidate Name", header_style),
            Paragraph("Register No", header_style),
            Paragraph("Demanded Heads", header_style),
            Paragraph("Amount Paid", header_style),
            Paragraph("Penalty Paid", header_style),
            Paragraph("Mode", header_style),
            Paragraph("Reference No", header_style)
        ]
        
        data = [headers]
        grouped_audit = group_payments_by_receipt(ledger_payments)
        for p in grouped_audit:
            data.append([
                Paragraph(f"#{p['receipt_number']}", cell_style),
                Paragraph(p['payment_date'].strftime('%Y-%m-%d') if p['payment_date'] else '-', cell_style),
                Paragraph(p['admission'].application.display_name, cell_style),
                Paragraph(p['admission'].registration_id or '-', cell_style),
                Paragraph(p['fee_types_str'], cell_style),
                Paragraph(f"₹{p['total_amount_paid']:.2f}", cell_style),
                Paragraph(f"₹{p['total_fine_paid']:.2f}", cell_style),
                Paragraph(p['payment_mode_display'], cell_style),
                Paragraph(p['reference_no'] or '-', cell_style)
            ])
            
        t = Table(data, colWidths=[60, 65, 120, 80, 100, 75, 75, 65, 100])
    else:
        headers = [
            Paragraph("Register No", header_style),
            Paragraph("Student Name", header_style),
            Paragraph("Term & Class", header_style),
            Paragraph("Fee Category", header_style),
            Paragraph("Net Demand", header_style),
            Paragraph("Paid Fees", header_style),
            Paragraph("Fines Collected", header_style),
            Paragraph("Pending Balance", header_style)
        ]
        
        data = [headers]
        for adm in admissions:
            active_class = adm.assigned_class or (adm.assigned_class_year.class_obj if adm.assigned_class_year else None)
            if active_class and adm.assigned_class_year and adm.assigned_fee_category:
                structure = FeeStructure.objects.filter(
                    academic_year=adm.application.academic_year,
                    institute=institute,
                    course=adm.selected_course,
                    class_obj=active_class,
                    class_year=adm.assigned_class_year,
                    fee_category=adm.assigned_fee_category
                ).first()
                
                if structure:
                    heads = structure.heads.filter(is_active=True)
                    if fee_type_id:
                        heads = heads.filter(fee_type_id=fee_type_id)
                        
                    std_demand = 0.0
                    std_collected = 0.0
                    std_fines = 0.0
                    
                    for head in heads:
                        if head.fee_type.is_discountable:
                            discount_pct = adm.custom_discount_percentage if adm.custom_discount_percentage is not None else adm.assigned_fee_category.discount_percentage
                            discount_pct = discount_pct or 0
                        else:
                            discount_pct = 0
                            
                        discounted_amount = float(head.amount) * (1 - float(discount_pct) / 100)
                        std_demand += discounted_amount
                        
                        payments = StudentFeePayment.objects.filter(admission=adm, fee_head=head, is_cancelled=False)
                        std_collected += sum(float(p.amount_paid) for p in payments)
                        std_fines += sum(float(p.fine_paid) for p in payments)
                        
                    pending_fee = std_demand - std_collected
                    pending_fee = max(0.0, pending_fee)
                    
                    if report_type == 'roster' and pending_fee <= 0:
                        continue
                        
                    class_term_str = f"{adm.assigned_class_year.name} - {adm.assigned_class.name}"
                    data.append([
                        Paragraph(adm.registration_id or '-', cell_style),
                        Paragraph(adm.application.display_name, cell_style),
                        Paragraph(class_term_str, cell_style),
                        Paragraph(adm.assigned_fee_category.name, cell_style),
                        Paragraph(f"₹{std_demand:.2f}", cell_style),
                        Paragraph(f"₹{std_collected:.2f}", cell_style),
                        Paragraph(f"₹{std_fines:.2f}", cell_style),
                        Paragraph(f"₹{pending_fee:.2f}", cell_style)
                    ])
                    
        t = Table(data, colWidths=[90, 140, 130, 110, 65, 65, 65, 75])
        
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)
    
    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Fee_Report_{report_type}_{datetime.datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def export_payments_pdf(request):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    
    institute = getattr(request.user, 'institute', None)
    if not institute:
        return HttpResponse("Unauthorized", status=401)
        
    payments = Payment.objects.filter(application__institute=institute).select_related('application__student')
    
    search_query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status', 'all')
    
    if search_query:
        payments = payments.filter(
            Q(application__student__first_name__icontains=search_query) |
            Q(application__student__username__icontains=search_query) |
            Q(gateway_transaction_id__icontains=search_query) |
            Q(application__id__icontains=search_query)
        ).distinct()

    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)
    if status_filter and status_filter != 'all':
        payments = payments.filter(status=status_filter)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'], fontSize=16, fontName='Helvetica-Bold', leading=20, alignment=1, spaceAfter=15)
    header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', textColor=colors.whitesmoke)
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, fontName='Helvetica', leading=10)
    
    elements.append(Paragraph(f"{institute.name.upper()}<br/><font size='12'>PAYMENT TRANSACTION REPORT</font>", title_style))
    elements.append(Spacer(1, 10))
    
    headers = [
        Paragraph("Form Number", header_style),
        Paragraph("Student Name", header_style),
        Paragraph("Mobile No.", header_style),
        Paragraph("Payment Status", header_style),
        Paragraph("Amount", header_style),
        Paragraph("Payment Mode", header_style),
        Paragraph("Gateway Transaction ID", header_style),
        Paragraph("Payment Date", header_style)
    ]
    
    data = [headers]
    for p in payments:
        student_name = p.application.display_name
        student_mobile = p.application.student_mobile
        created_at_str = p.payment_date.strftime('%Y-%m-%d') if p.payment_date else (p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '')
        status_display = dict(Payment._meta.get_field('status').choices).get(p.status, p.status).title()
        
        data.append([
            Paragraph(f"#{p.application.id}", cell_style),
            Paragraph(student_name, cell_style),
            Paragraph(student_mobile or '-', cell_style),
            Paragraph(status_display, cell_style),
            Paragraph(f"₹{p.amount:.2f}", cell_style),
            Paragraph(p.payment_mode or '-', cell_style),
            Paragraph(p.gateway_transaction_id or '-', cell_style),
            Paragraph(created_at_str, cell_style)
        ])
        
    t = Table(data, colWidths=[50, 140, 90, 80, 70, 70, 130, 110])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)
    
    doc.build(elements)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Payment_Details_{institute.name}.pdf"'
    return response


@login_required
def activity_logs_view(request):
    if request.user.role != 'institute' and not request.user.is_staff and not request.user.is_superuser:
        return redirect('/')
        
    institute = getattr(request.user, 'institute', None)
    if not institute and (request.user.is_staff or request.user.is_superuser):
        institute = Institute.objects.first()

    if not institute:
        messages.error(request, "No Institute context found.")
        return redirect('/')

    selected_user_id = request.GET.get('user_id')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    from .models import UserActivityLog
    logs_qs = UserActivityLog.objects.filter(institute=institute).select_related('user').order_by('-created_at')
    
    if selected_user_id:
        logs_qs = logs_qs.filter(user_id=selected_user_id)
    if date_from:
        logs_qs = logs_qs.filter(created_at__date__gte=date_from)
    if date_to:
        logs_qs = logs_qs.filter(created_at__date__lte=date_to)
        
    from django.core.paginator import Paginator
    paginator = Paginator(logs_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(activity_logs__institute=institute).distinct()
    
    if not users.exists():
        from django.db.models import Q
        users = User.objects.filter(Q(role='institute') | Q(is_staff=True) | Q(is_superuser=True))
        
    return render(request, 'institute/activity_logs.html', {
        'page_obj': page_obj,
        'users': users,
        'selected_user_id': selected_user_id,
        'date_from': date_from,
        'date_to': date_to,
    })


@login_required
def manage_attendance(request):
    institute = get_current_institute(request)
    if not institute:
        messages.error(request, "No Institute context found.")
        return redirect('/')

    from academics.models import Class, StudentAttendance
    classes = Class.objects.filter(institute=institute).select_related('course')
    
    selected_class_id = request.GET.get('class_id')
    selected_date_str = request.GET.get('date', datetime.date.today().strftime('%Y-%m-%d'))
    
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except Exception:
        selected_date = datetime.date.today()
        selected_date_str = selected_date.strftime('%Y-%m-%d')

    selected_class = None
    admissions = []
    
    if selected_class_id:
        selected_class = Class.objects.filter(id=selected_class_id).first()
        if selected_class:
            admissions = list(Admission.objects.filter(
                Q(assigned_class=selected_class) | Q(assigned_class__isnull=True, selected_course=selected_class.course)
            ).exclude(status='trashed').select_related('application__student'))

    if request.method == 'POST' and selected_class:
        att_date_str = request.POST.get('attendance_date', selected_date_str)
        try:
            att_date = datetime.datetime.strptime(att_date_str, '%Y-%m-%d').date()
        except Exception:
            att_date = datetime.date.today()

        for adm in admissions:
            status_val = request.POST.get(f'status_{adm.id}', 'present')
            remarks_val = request.POST.get(f'remarks_{adm.id}', '')
            StudentAttendance.objects.update_or_create(
                admission=adm,
                date=att_date,
                defaults={
                    'status': status_val,
                    'remarks': remarks_val,
                    'marked_by': request.user
                }
            )
        messages.success(request, f"Attendance for {selected_class.name} on {att_date} saved successfully!")
        return redirect(f"{request.path}?class_id={selected_class.id}&date={att_date_str}")

    # Fetch existing attendance logs for the selected date
    existing_att = {}
    if selected_class and selected_date:
        logs = StudentAttendance.objects.filter(
            admission__assigned_class=selected_class,
            date=selected_date
        )
        for log in logs:
            existing_att[log.admission_id] = log

    student_rows = []
    for adm in admissions:
        log = existing_att.get(adm.id)
        student_rows.append({
            'admission': adm,
            'status': log.status if log else 'present',
            'remarks': log.remarks if log else ''
        })

    return render(request, 'institute/manage_attendance.html', {
        'classes': classes,
        'selected_class': selected_class,
        'selected_class_id': selected_class_id,
        'selected_date_str': selected_date_str,
        'student_rows': student_rows
    })


# =========================
# TEACHER STUDENT DOCUMENT MANAGEMENT
# =========================

@login_required
def upload_student_document_by_teacher(request, admission_id):
    """
    Allows institute staff/teachers to upload internal student documents
    (e.g., Student ID Card, Previous Marksheet, Transfer Certificate, etc.)
    Hidden from student portal login.
    """
    if request.user.role != 'institute':
        messages.error(request, "Access denied.")
        return redirect('institute_root')

    admission = get_object_or_404(Admission, id=admission_id, application__institute=request.user.institute)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        doc_type = request.POST.get('doc_type', 'Other').strip()
        doc_file = request.FILES.get('document_file')

        if not title or not doc_file:
            messages.error(request, "Please provide both document title and file.")
        else:
            doc = StudentDocument.objects.create(
                admission=admission,
                title=title,
                doc_type=doc_type,
                file=doc_file,
                uploaded_by=request.user,
                is_teacher_uploaded=True
            )
            messages.success(request, f"Document '{doc.title}' ({doc.doc_type}) uploaded successfully for {admission.application.display_name}!")

    return redirect(request.META.get('HTTP_REFERER', 'student_list'))


@login_required
def delete_student_document_by_teacher(request, doc_id):
    """
    Allows teachers/staff to delete an uploaded student document.
    """
    if request.user.role != 'institute':
        messages.error(request, "Access denied.")
        return redirect('institute_root')

    doc = get_object_or_404(StudentDocument, id=doc_id, admission__application__institute=request.user.institute)
    title = doc.title
    if doc.file:
        doc.file.delete(save=False)
    doc.delete()
    messages.success(request, f"Document '{title}' deleted successfully.")

    return redirect(request.META.get('HTTP_REFERER', 'student_list'))


# =========================
# HELPER & MULTI-INSTITUTE SWITCHER
# =========================

def get_current_institute(request):
    """
    Helper function to retrieve the active institute for the logged-in user.
    Falls back to request.user.institute if not set.
    """
    if hasattr(request.user, 'get_active_institute'):
        inst = request.user.get_active_institute(request)
        if inst:
            return inst
    return getattr(request.user, 'institute', None)


@login_required
def switch_active_institute(request, institute_id):
    """
    Switches the active institute context for multi-tagged staff/admin users.
    """
    from .models import Institute
    target_inst = get_object_or_404(Institute, id=institute_id)
    accessible = request.user.get_accessible_institutes() if hasattr(request.user, 'get_accessible_institutes') else [request.user.institute]

    if request.user.is_superuser or target_inst in accessible:
        request.session['active_institute_id'] = target_inst.id
        messages.success(request, f"Switched active institute to '{target_inst.name}'")
    else:
        messages.error(request, "You do not have access to that institute.")

    return redirect(request.META.get('HTTP_REFERER', '/institute/dashboard/'))


# =========================
# EMPLOYEE PRIVILEGES & MULTI-INSTITUTE TAGGING MANAGEMENT
# =========================

@login_required
def employee_privileges_view(request):
    """
    Admin control panel to tag institutes and set module privileges for staff/teachers.
    """
    if not (request.user.is_superuser or request.user.is_staff or request.user.has_privilege('perm_employee_privileges')):
        messages.error(request, "Access denied. Employee Privileges Management requires administrative rights.")
        return redirect('institute_root')

    from accounts.models import User, UserPrivilege
    from .models import Institute

    # Search & Filter Users (Only staff/institute role users)
    user_search = request.GET.get('user_search', '').strip()

    users_qs = User.objects.filter(Q(role='institute') | Q(is_staff=True) | Q(is_superuser=True)).order_by('username')
    if user_search:
        users_qs = users_qs.filter(
            Q(username__icontains=user_search) |
            Q(first_name__icontains=user_search) |
            Q(last_name__icontains=user_search) |
            Q(email__icontains=user_search)
        )

    selected_user_id = request.GET.get('user_id') or request.POST.get('user_id')
    selected_user = None
    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()
    if not selected_user and users_qs.exists():
        selected_user = users_qs.first()

    all_institutes = Institute.objects.all().order_by('name')

    # Fetch or create UserPrivilege object
    user_priv = None
    if selected_user:
        user_priv, _ = UserPrivilege.objects.get_or_create(user=selected_user)

    if request.method == 'POST' and selected_user:
        # 1. Update Primary Default Institute
        default_inst_id = request.POST.get('default_institute')
        if default_inst_id:
            default_inst = Institute.objects.filter(id=default_inst_id).first()
            if default_inst:
                selected_user.institute = default_inst
                selected_user.save()

        # 2. Update Tagged Multi-Institutes
        tagged_inst_ids = request.POST.getlist('tagged_institutes')
        tagged_institutes = Institute.objects.filter(id__in=tagged_inst_ids)
        selected_user.accessible_institutes.set(tagged_institutes)

        # 3. Update Module Privileges
        perm_fields = [
            'perm_admissions_overview', 'perm_student_registration', 'perm_student_list', 'perm_rank_list',
            'perm_attendance', 'perm_notices', 'perm_timetables', 'perm_academic_results', 'perm_course_inventory',
            'perm_fee_reports', 'perm_fee_receipts', 'perm_payment_details',
            'perm_activity_logs', 'perm_system_backup', 'perm_employee_privileges'
        ]

        for p_field in perm_fields:
            setattr(user_priv, p_field, request.POST.get(p_field) == 'on')
        user_priv.save()

        messages.success(request, f"Employee Privileges and Tagged Institutes updated successfully for {selected_user.username}!")
        return redirect(f"{request.path}?user_id={selected_user.id}")

    # Tagged institutes for selected user
    user_tagged_inst_ids = list(selected_user.accessible_institutes.values_list('id', flat=True)) if selected_user else []

    return render(request, 'institute/employee_privileges.html', {
        'users': users_qs,
        'selected_user': selected_user,
        'user_priv': user_priv,
        'all_institutes': all_institutes,
        'user_tagged_inst_ids': user_tagged_inst_ids,
        'user_search': user_search,
    })



